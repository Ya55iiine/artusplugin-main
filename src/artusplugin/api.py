# -*- coding: utf-8 -*-
#
# Copyright (C) 2016 Artus
# All rights reserved.
#
# Author: Michel Guillot <michel.guillot@meggitt.com>

""" Central functionality for the Artus plugin """

# Genshi
# from genshi.builder import tag
from trac.util.html import html as tag

# Trac
from trac.attachment import IAttachmentChangeListener, Attachment
from trac.config import OrderedExtensionsOption
from trac.core import Component as CComponent, implements, TracError
from trac.perm import IPermissionRequestor
from trac.resource import ResourceNotFound, ResourceSystem, Resource, \
    get_resource_url, get_resource_name
from trac.ticket import Ticket
from trac.ticket.api import ITicketManipulator, ITicketChangeListener, ITicketActionController
from trac.ticket.model import Component, Version
from trac.util.datefmt import utc, localtz
from trac.util.text import exception_to_unicode, pretty_size
from trac.versioncontrol.api import RepositoryManager
from trac.wiki.api import IWikiSyntaxProvider

# Standard lib
from datetime import datetime
# from ldap_utilities import Ldap_Utilities
from artusplugin.ldap.ldap_utilities import Ldap_Utilities
import filecmp
import glob
import os
import re
import shutil
import sys
import syslog
import warnings
from itertools import chain

# https://github.com/JoshData/python-email-validator
from email_validator import validate_email, EmailNotValidError

# Tracchildtickets
from childtickets import TracchildticketsModule

# MasterTickets
from mastertickets.api import IMasterObserver, MasterTicketsSystem
from mastertickets.util import linkify_ids

# Same package
from artusplugin import util, form, model, web_ui, cache, admin, _
from artusplugin.advanced_workflow import TicketWF


_UNCHECKED = None
_CHECKED = 'checked'


class ArtusSystem(CComponent):
    """Central functionality for the Artus plugin."""

    implements(IWikiSyntaxProvider,
               ITicketManipulator,
               IMasterObserver,
               ITicketChangeListener,
               IAttachmentChangeListener,
               IPermissionRequestor)

    action_controllers = OrderedExtensionsOption('ticket', 'workflow',
                                                 ITicketActionController,
                                                 default='ConfigurableTicketWorkflow',
                                                 include_missing=False,
                                                 doc="""Ordered list of workflow controllers to use for ticket actions (''since 0.11'').""")

    def __init__(self):
        CComponent.__init__(self)
        program_data = util.get_program_data(self.env)
        self.trac_env_name = program_data['trac_env_name']
        self.program_name = program_data['program_name']
        self._skills = {'qa_review': ['DIR', 'PROC-1', 'PROC-2', 'PROC-3', 'PROC-4',
                                      'PROC-5', 'PROC-6', 'PROC-7', 'PROC-8-1',
                                      'PROC-8-2', 'PROC-8-3', 'PROC-8-4', 'PROC-9',
                                      'PROC-10', 'PROC-11', 'PROC-12', 'PROC-13'],
                        'qa_audit': ['QMS', 'CLI', 'INT']}
        self._skills['proj'] = util.get_prop_values(self.env, 'skill_dirs').keys()

    # IWikiSyntaxProvider methods

    def get_wiki_syntax(self):
        programidre = self.env.config.get('artusplugin', 'programidre')
        skill_options = self.env.config.get('ticket-custom', 'skill.options')
        unmanaged_skills = self.env.config.get('artusplugin', 'unmanaged_skills', 'EXT')
        suffx = util.get_prop_values(self.env, 'attachment_edit.office_suite')

        repository_regexp = r'(?:/\w+)?/(?:trunk|tags|branches)/.+'
        yield (repository_regexp, self._format_repository_link)

        archive_name_regexp = '%s/ECM_%s_\d{3}(?:\.\d+)?.zip' % (self.trac_env_name, programidre)
        yield (archive_name_regexp, self._format_archive_name_link)

        version_tag_regexp_list = []
        component_MA_regexp = self._regexp_group_strip(
            model.NamingRule.get_component_pattern(programidre, skill_options, 'MA'))
        component_SER_regexp = self._regexp_group_strip(
            model.NamingRule.get_component_pattern(programidre, skill_options, 'SER'))
        document_UM_regexp = self._regexp_group_strip(
            model.NamingRule.get_document_pattern(programidre, skill_options, 'UM', unmanaged_skills))
        document_SER_regexp = self._regexp_group_strip(
            model.NamingRule.get_document_pattern(programidre, skill_options, 'SER'))
        document_ER_regexp = self._regexp_group_strip(
            model.NamingRule.get_document_pattern(programidre, skill_options, 'ER'))
        ecm_regexp = self._regexp_group_strip(
            model.NamingRule.get_ecm_pattern(programidre))
        fee_regexp = self._regexp_group_strip(
            model.NamingRule.get_fee_pattern(programidre))
        version_tag_regexp_list.extend([component_MA_regexp, component_SER_regexp,
                                        document_UM_regexp, document_SER_regexp,
                                        document_ER_regexp, ecm_regexp, fee_regexp])
        version_tag_regexp_list.append(model.NamingRule.get_version_pattern(
            programidre, skill_options, 'component', 'MA'))
        version_tag_regexp_list.append(model.NamingRule.get_version_pattern(
            programidre, skill_options, 'component', 'SER'))
        version_tag_regexp = '(?:' + '|'.join(version_tag_regexp_list) + ')'
        archive_file_regexp = '%s/(?:[^/]+/)*[^/]+$' % version_tag_regexp
        yield (archive_file_regexp, self._format_archive_file_link)
        
        archive_dir_regexp = '%s/(?:[^/]+/)*$' % version_tag_regexp
        yield (archive_dir_regexp, self._format_archive_dir_link)
        
        docfile_regexp_list = []
        docfile_regexp_list.append(r'^(?!\[attachment:)[^/]*?[^":/]+\.[pP][dD][fF]')
        docfile_regexp_list.append(str('^(?!\[attachment:)[^/]*?[^":]+\.(?:%s)' % '|'.join(suffx.keys())))
        docfile_regexp = '(?:' + '|'.join(docfile_regexp_list) + ')'
        yield (docfile_regexp, self._format_docfile_link)

        masterticket_regexp = r"^\s*\d+\s*(?:,\s*\d+\s*)*$"
        yield (masterticket_regexp, self._format_masterticket_link)
                                
        milestone_tag_regexp = self._regexp_group_strip(
            model.NamingRule.get_milestone_pattern(programidre, skill_options))
        # For legacy milestones with '_' instead of '-'
        milestone_tag_regexp = milestone_tag_regexp.replace('[A-Za-z0-9-.]',
                                                            '[A-Za-z0-9-._]')
        yield ( milestone_tag_regexp, self._format_milestone_tag_link)

        yield (version_tag_regexp, self._format_version_tag_link)
                
        version_regexp = model.NamingRule.get_version_pattern(
            programidre, skill_options, 'document', 'ER')
        yield (version_regexp, self._format_version_link)
                
        ticket_regexp_list = []
        ticket_regexp_list.append('EFR_%s_(?:(?:%s)_)?\d{3}' % (programidre, skill_options))
        ticket_regexp_list.append('(?:ECR|MEMO)_%s_(?:%s)_\d{3}' % (programidre, skill_options))
        ticket_regexp_list.append('ECM_%s_\d{3}(?:_v\d{1,2})?' % programidre)
        ticket_regexp_list.append('DOC_%s' % model.NamingRule.get_version_pattern(
            programidre, skill_options, 'document', 'ER'))
        ticket_regexp_list.append('P?RF_%s_[a-z-]+\.?[a-z-]+' % component_MA_regexp)
        ticket_regexp_list.append('P?RF_%s_[a-z-]+\.?[a-z-]+' % component_SER_regexp)
        ticket_regexp_list.append('P?RF_%s_[a-z-]+\.?[a-z-]+' % document_UM_regexp)
        ticket_regexp_list.append('P?RF_%s_[a-z-]+\.?[a-z-]+' % document_SER_regexp)
        ticket_regexp_list.append('P?RF_%s_[a-z-]+\.?[a-z-]+' % document_ER_regexp)
        ticket_regexp_list.append('P?RF_%s_[a-z-]+\.?[a-z-]+' % ecm_regexp)
        ticket_regexp_list.append('P?RF_%s_[a-z-]+\.?[a-z-]+' % fee_regexp)
        milestoneshort_regexp = '[A-Za-z0-9-._]+'
        ticket_regexp_list.append('MOM_%s_(?:%s)_PMC_%s' % (
            programidre, skill_options, milestoneshort_regexp))
        milestonetagshort_regexp = milestoneshort_regexp + '\.(?:Prepared|Reviewed|Accepted)\d*'
        ticket_regexp_list.append('MOM_%s_(?:%s)_(?:CCB|Review)_%s' % (
            programidre, skill_options, milestonetagshort_regexp))
        chrono_regexp = '\d{2}-\d{3}'
        ticket_regexp_list.append('MOM_%s_(?:%s)_(?:AUD|REV)_%s' % (
            programidre, skill_options, chrono_regexp))
        ticket_regexp_list.append('(?:AI|RISK)_%s_(?:%s)_\d{4}' % (
            programidre, skill_options))
        ticket_regexp = '(?:' + '|'.join(ticket_regexp_list) + ')'
        yield (ticket_regexp, self._format_ticket_link)
               
        ci_regexp = model.NamingRule.get_ci_pattern(programidre, skill_options)
        yield (ci_regexp, self._format_ci_link)

    def _regexp_group_strip(self, regexp):
        return re.sub(r'\(\?P<\w+>(.+?)\)', r'(?:\1)', regexp)

    def _format_repository_link(self, formatter, ns, match):
        idx = ns.find("?rev=")
        if idx == -1:
            idx = ns.find(_(" (on behalf of"))
        else:
            m = re.search(r'\?rev=(\d+)', ns)
            rev = m.group(1)
            idx = idx + len("?rev=") + len(rev)

        if idx == -1:
            path = ns
            remainder = ""
        else:
            path = ns[:idx]
            remainder = ns[idx:]

        return tag.a(path, href=formatter.href.browser(util.get_url(path),
                     rev=util.get_revision(path))) + remainder

    def _format_docfile_link(self, formatter, ns, match):
        resource = formatter.context.resource
        if resource.realm == 'ticket':
            ticket = Ticket(self.env, resource.id)
            if ticket['type'] in ('ECM', 'FEE', 'DOC', 'PRF'):
                repo_url = (ticket['sourceurl'] if ticket['type'] in ('ECM', 'FEE', 'DOC')
                            else ticket['documenturl'])
                url = util.get_url(repo_url)
                revision = util.get_revision(repo_url)
                # Opening in new instance via ClickOnce
                # for Basic / Kerberos peaceful coexistence
                # See trac.conf rewrite rules for MS and PDF files
                # Also avoids  Internet Explorer Open or Save Popup
                # See: https://jwcooney.com/2014/03/31/remove-internet-explorer-open-or-save-popup/
                url = util.get_tracbrowserurl(self.env, url)
                href_val = '%s/%s?format=raw&rev=%s' % (url, ns, revision)
                target = '_self'
                docfile_link = "pdffile_link" if ns.endswith('.pdf') else "srcfile_link"
                return tag.a(ns,
                             href=href_val,
                             target=target,
                             title="View File (document folder in the repository on the Trac server)",
                             id=docfile_link)
            else:
                return ns
        elif resource.realm == 'changeset':
            try:
                skill = util.get_skill(self.env, ns, self.program_name)
                rm = RepositoryManager(self.env)
                rn = util.get_repo_name(self.env, skill)
                repos = rm.get_repository(rn)
                changeset = repos.get_changeset(resource.id)
                match = re.search('ticket:(\d+)', changeset.message)
                if match:
                    ticket = Ticket(self.env, match.group(1))
                    if ticket['type'] in ('ECM', 'FEE', 'DOC'):
                        sourceurl = ticket['sourceurl']
                        return tag.a(ns, href=formatter.href.browser(
                            '%s/%s' % (util.get_url(sourceurl), ns),
                            format='raw',
                            rev=util.get_revision(sourceurl)),
                            title="View File (document folder in the repository on the Trac server)")
            except Exception:
                return ns
        else:
            return ns

    def _format_masterticket_link(self, formatter, ns, match):
        if formatter.context.resource.realm == 'ticket':
            return linkify_ids(self.env, formatter.req,
                               set([lid.strip() for lid in ns.split(',')]))
        else:
            return ns

    def _format_ticket_link(self, formatter, ns, match):
        # Notes:
        # * Do not use quotes around % sign or it will be escaped
        # * Do not forget the comma after the parameter or it will not be a tuple
        for tid in self.env.db_query("""
                SELECT id FROM ticket
                WHERE summary=%s
                """, (ns,)):
            break
        try:
            tid
            return tag.a(ns, href=formatter.href.ticket('%s' % tid),
                         title='Ticket #%s' % tid)
        except NameError:
            return ns

    def _format_milestone_tag_link(self, formatter, ns, match):
        try:
            model.Tag(self.env, ns)
            return tag.a(ns, href=formatter.href.admin('tags_mgmt/milestone_tags/%s' %
                                                       ns), title=ns)
        except ResourceNotFound:
            return ns

    def _format_version_tag_link(self, formatter, ns, match):
        try:
            model.Tag(self.env, ns)
            return tag.a(ns, href=formatter.href.admin('tags_mgmt/version_tags/%s' %
                                                       ns), title=ns)
        except ResourceNotFound:
            return ns

    def _format_archive_name_link(self, formatter, ns, match):
        zip_name = ns.split('/')[1]
        ticket = Ticket(self.env, formatter.context.resource.id)
        for archive_path in cache.PDFPackage.get_archives_paths(ticket):
            if os.path.basename(archive_path) == zip_name:
                fileSize = os.path.getsize(archive_path)
                break
        else:
            # Archive not found
            return tag.div(zip_name, ':', style='margin-top:1em')
        return tag.div(tag.a(zip_name, href=formatter.href('PDF-packaging', zip_name), title=zip_name), ' (%s):' % pretty_size(fileSize), style='margin-top:1em')

    def _format_archive_dir_link(self, formatter, ns, match):
            vtag_name, path = ns.split('/', 1)
            try:
                vtag = model.Tag(self.env, name=vtag_name)
                tag_url = util.get_url(vtag.tag_url)
            except ResourceNotFound:
                tag_url = None
            if path:
                dirname = path.rstrip('/').rsplit('/', 1)[-1]
                tabs = 1 + len(path.rstrip('/').split('/'))
            else:
                dirname = vtag_name
                tabs = 1
            return tag.span(tag.a(dirname, href=formatter.href.browser(tag_url, path), title=dirname) if tag_url else dirname, ':', style='margin-left:%sem' % tabs)
            return ns        

    def _format_archive_file_link(self, formatter, ns, match):
            ticket = Ticket(self.env, formatter.context.resource.id)
            archives_content = cache.PDFPackage.get_archives_content(ticket)
            archives_documents = cache.PDFPackage.get_archives_documents(archives_content)
            vtag_name, path = ns.split('/', 1)
            if vtag_name in archives_documents and path not in archives_documents[vtag_name]:
                # renamed
                archive_path = "%s.pdf" % vtag_name
            else:
                # not renamed
                archive_path = path
            try:
                vtag = model.Tag(self.env, name=vtag_name)
                tag_url = util.get_url(vtag.tag_url)
            except ResourceNotFound:
                tag_url = None
            if '/' in archive_path:
                filename = path.rsplit('/', 1)[-1]
                tabs = 1 + len(path.split('/'))
            else:
                filename = archive_path
                tabs = 2
            return tag.span(tag.a(filename, href=formatter.href.browser(tag_url, path), title=filename) if tag_url else filename, style='margin-left:%sem' % tabs)


    def _format_version_link(self, formatter, ns, match):
        try:
            Version(self.env, ns)
            return tag.a(ns, href=formatter.href.admin('tags_mgmt/versions/%s' %
                                                       ns), title=ns)
        except ResourceNotFound:
            return ns

    def _format_ci_link(self, formatter, ns, match):
        try:
            model.Document(self.env, ns)
            ci_group = 'documents'
        except ResourceNotFound:
            try:
                Component(self.env, ns)
                ci_group = 'components'
            except ResourceNotFound:
                return ns
        return tag.a(ns, href=formatter.href.admin('tags_mgmt/%s/%s' %
                                                   (ci_group, ns)), title=ns)

    def get_link_resolvers(self):
        yield ('attachment', self._format_link)

    def _format_link(self, formatter, ns, target, label):

        link = target
        ids = link.split(':', 2)
        attachment = None
        if len(ids) == 3:
            known_realms = ResourceSystem(self.env).get_known_realms()
            # new-style attachment: TracLinks (filename:realm:id)
            if ids[1] in known_realms:
                attachment = Resource(ids[1], ids[2]).child('attachment',
                                                            ids[0])
            else:  # try old-style attachment: TracLinks (realm:id:filename)
                if ids[0] in known_realms:
                    attachment = Resource(ids[0], ids[1]).child('attachment',
                                                                ids[2])
        else:  # local attachment: TracLinks (filename)
            attachment = formatter.resource.child('attachment', link)
        if attachment and 'ATTACHMENT_VIEW' in formatter.perm(attachment):
            try:
                model = Attachment(self.env, attachment)
                raw_href = get_resource_url(self.env, attachment,
                                            formatter.href, format='raw')
                if ns.startswith('raw'):
                    return tag.a(label, class_='attachment',
                                 href=raw_href,
                                 title=get_resource_name(self.env, attachment))
                href = get_resource_url(self.env, attachment, formatter.href)
                title = get_resource_name(self.env, attachment)
                return tag(tag.a(label, class_='attachment', title=title,
                                 href=href),
                           tag.a(u'\u200b', class_='trac-rawlink',
                                 href=raw_href, title=_("Download")))
            except ResourceNotFound:
                pass
            # FIXME: should be either:
            #
            # model = Attachment(self.env, attachment)
            # if model.exists:
            #     ...
            #
            # or directly:
            #
            # if attachment.exists:
            #
            # (related to #4130)
        return tag.a(label, class_='missing attachment')

    # IPermissionRequestor method

    def get_permission_actions(self):
        """ Definition of rights regarding ticket management
        TICKET_FORCE_EDIT : right to force the edit mode
        ATTACHMENT_FORCE_EDIT : right to force the edit mode

                                authenticated      developer     authorized      admin
        TICKET_FORCE_EDIT                                                          X
        ATTACHMENT_FORCE_EDIT                                                      X

        """
        return ['TICKET_FORCE_EDIT', 'ATTACHMENT_FORCE_EDIT']

    # ITicketManipulator methods

    def prepare_ticket(self, req, ticket, fields, actions):
        """ Not currently called, but should be provided
            for future compatibility """
        pass

    def validate_ticket(self, req, ticket):
        """Validate a ticket after it's been populated from user input.

        Must return a list of `(field, message)` tuples, one for each problem
        detected. `field` can be `None` to indicate an overall problem with the
        ticket. Therefore, a return value of `[]` means everything is OK.

        Note that 'field' MUST be 'None' if message is a markup stream
        built with Genshi 'tag' object """

        # Rights management
        issues = []

        # Don't lose time to validate for ticket preview
        # TDB: what is the purpose of doing so ?
        if 'x-requested-with' in dict(req._inheaders).keys():
            return issues

        action = req.args.get('action')
        if ticket.exists and action is None:
            # May happen if no action is selected when changes are submitted
            # This is the case if all actions are disabled
            return issues

        # ------------------
        # Ticket in creation
        # ------------------

        if not ticket.exists:
            # Check unicity for tickets WITHOUT chronological number
            if (ticket['type'] in ('RF', 'PRF', 'DOC') or
                (ticket['type'] == 'MOM' and
                 (ticket['momtype'] in ('CCB', 'Progress') or
                  (ticket['momtype'] == 'Review' and
                   ticket['skill'] in self._skills['proj'])))):

                if ticket['type'] == 'MOM':
                    if ticket['momtype'] == 'CCB':
                        if ticket['milestonetag']:
                            milestonetag_short = ticket['milestonetag'].rsplit(
                                '_', 1)[-1]
                            ticket['summary'] = '%s_%s_%s_%s_%s' % (
                                ticket['type'],
                                self.program_name,
                                ticket['skill'],
                                ticket['momtype'],
                                milestonetag_short)
                        else:
                            issues.append((None,
                                           'The milestonetag field is required.'))
                    elif ticket['momtype'] == 'Progress':
                        if ticket['milestone']:
                                milestone_short = ticket['milestone'].rsplit(
                                    '_', 1)[-1]
                                ticket['summary'] = '%s_%s_%s_%s_%s' % (
                                                    ticket['type'],
                                                    self.program_name,
                                                    ticket['skill'],
                                                    'PMC',
                                                    milestone_short)
                        else:
                            issues.append((
                                None,
                                'The milestone field is required.'))
                    elif ticket['momtype'] == 'Review':
                        if ticket['skill'] in self._skills['proj']:
                            if ticket['milestonetag']:
                                milestonetag_short = ticket['milestonetag'].rsplit(
                                    '_', 1)[-1]
                                ticket['summary'] = '%s_%s_%s_%s_%s' % (
                                                    ticket['type'],
                                                    self.program_name,
                                                    ticket['skill'],
                                                    ticket['momtype'],
                                                    milestonetag_short)
                            else:
                                issues.append((
                                    None,
                                    'The milestonetag field is required.'))

                db = self.env.get_db_cnx()
                cursor = db.cursor()
                cursor.execute("SELECT id FROM ticket "
                               "WHERE summary=%s", (ticket['summary'],))
                row = cursor.fetchone()
                if row:
                    msg = tag.p('The ticket %s already exists as ticket ' %
                                ticket['summary'],
                                tag.a("#%s" % row[0],
                                      href="%s" % req.href.ticket(row[0])))
                    issues.append((None, msg))

            # Fill in milestone field from milestonetag field
            if ticket['type'] == 'MOM' and ticket['momtype'] in ('CCB', 'Review'):
                tg = model.Tag(self.env, ticket['milestonetag'])
                ticket['milestone'] = tg.tagged_item

            # Set ticket status at creation
            ticket['status'] = TicketWF.get_WF(ticket).get_initial_status()

            tickets_with_forms = set(('EFR', 'ECR', 'RF', 'PRF'))

            if 'MOM' in util.get_prop_values(self.env,
                                             'ticket_edit.office_suite').keys():
                tickets_with_forms.add('MOM')

            if ticket['type'] in tickets_with_forms:
                # Set skill
                skill = None
                skills = self.env.config.get('ticket-custom', 'skill.options')
                if ticket['type'] == 'EFR':
                    skill = 'SYS' if 'SYS' in skills else ticket['skill']
                elif ticket['type'] in ('ECR', 'MOM'):
                    skill = ticket['skill']
                else:
                    # RF / PRF
                    document = ticket['document']  # This is the version tag !

                    # No check for PRF on documents for unmanaged skills or
                    # for PRF on coordination memos or evolution sheets:
                    # edition, revision and status are NOT extracted
                    # from document or memo reference and therefore
                    # are NOT copied by TRAC because of the unknown formalism
                    # only the document reference is copied by TRAC
                    # for unmanaged skills
                    if (not util.skill_is_unmanaged(self.env, document) and
                        not document.startswith('ECM_%s_' % self.program_name) and
                        not document.startswith('FEE_%s_' % self.program_name)):
                        regular_expression = r"\A%s_(%s)(?:_(?:\w|-)+)?_" % (
                            self.program_name,
                            self.env.config.get('ticket-custom',
                                                'skill.options'))
                        "(?:(?:\w|-)+)_[1-9]\d*\.(?:0|[1-9]\d*)\."
                        "(?:Draft[1-9]\d*|Proposed[1-9]\d*|Released)\Z"
                        match = re.search(regular_expression, document)
                        if not match:
                            msg = tag.p("Sorry, can not save your changes. "
                                        "The document tag is not well formed ! "
                                        "Here is the regular expression "
                                        "the tag should conform to: %s \n"
                                        % regular_expression)
                            msg += tag.p("See the following link "
                                         "for more details: ",
                                         tag.a("Document identification",
                                               href="%s/index.php?post/106" %
                                               self.env.config.get('artusplugin',
                                                                   'dc_url')))
                            issues.append((None, msg))
                        else:
                            skill = match.group(1)
                    else:
                        skill = self.env.config.get('artusplugin', 'default_skill')

                    if skill:
                        ticket['skill'] = skill

                    # Check the existence of the associated template
                    # when configurable by the user
                    program_path = '/var/cache/trac/tickets/%s' % self.trac_env_name
                    template_cls = form.TicketFormTemplate.get_subclass(self.env,
                                                                        ticket['type'])
                    if template_cls.configurable:
                        template = template_cls(self.env,
                                                ticket['type'],
                                                program_path,
                                                skill)
                        if template.repo_form:
                            repos = util.get_repository(self.env,
                                                        template.source_name)
                            repo_path = util.repo_path(template.source_name)
                            if not repos.has_node(repo_path):
                                msg = tag.p("The following form template "
                                            "was not found: %s (rev=%s)\n"
                                            % (template.source_name,
                                               template.source_rev))
                                msg += tag.p("Check the ticket form template path "
                                             "specification on the appropriate ",
                                             tag.a("admin panel",
                                                   href=req.href.admin('ticket',
                                                                       'template')))
                                issues.append((None, msg))

            # Check the baseline field is not empty, exists and is of the same skill than the ticket
            if ticket['type'] == 'ECR':
                if ticket['document']:
                    try:
                        model.Tag(self.env, ticket['document'])
                        tg_skill = util.get_skill(self.env, ticket['document'], self.program_name)
                        if tg_skill != 'EXT' and ticket['skill'] != tg_skill:
                            issues.append((None, 'The baseline tag %s has not the same skill as the ticket (%s).' %
                                           (ticket['document'], ticket['skill'])))
                    except ResourceNotFound:
                        issues.append((None, 'The baseline tag %s does not exist.' % ticket['document']))
                else:
                    issues.append((None, 'The baseline tag field is required.'))

            # Check validity of parent MOM
            if ticket['type'] in ('RF', 'PRF') and ticket['parent']:
                # Check parent ticket is DOC or ECM/FEE
                msg = None
                try:
                    tid = int(ticket.values.get('parent').lstrip('#'))
                    tkt = Ticket(self.env, tid)
                    if tkt['type'] not in ('ECM', 'FEE', 'DOC'):
                        msg = "Ticket %s is not an ECM or FEE or DOC ticket" % tid
                    if tkt['type'] == 'DOC' and tkt['skill'] != ticket['skill']:
                        msg = "Ticket %s is not a DOC for skill %s" % (
                            tid,
                            ticket['skill'])
                except Exception:
                    e = sys.exc_info()[1]
                    try:
                        msg = exception_to_unicode(e)
                    finally:
                        del e
                if msg:
                    label = web_ui.Ticket_UI.field_label(ticket, 'parent')
                    issues.append((label, msg))

            # Reserve space for description field to host TicketQueryMacro
            if ((ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy and
                 ticket['ecmtype'] == 'Technical Note') or
                ticket['type'] == 'FEE' or
                ticket['type'] == 'DOC'):
                ticket['description'] = '[[br]]'

            # Checks previous version closed before creating a new one
            if ticket['type'] == 'DOC':
                ci_name = ticket['configurationitem']

                # Check sourceurl not the same as other open tickets on same CI
                db = self.env.get_db_cnx()
                cursor = db.cursor()
                cursor.execute("SELECT id FROM ticket t,ticket_custom tc "
                               "WHERE t.type='DOC' "
                               "AND t.status <> 'closed' "
                               "AND t.id=tc.ticket "
                               "AND tc.name='configurationitem' "
                               "AND tc.value='%s'" % ci_name)
                other_tkts = []
                for row in cursor:
                    tkt = Ticket(self.env, row[0])
                    if util.get_url(tkt['sourceurl']) == util.get_url(ticket['sourceurl']):
                        other_tkts.append((str(tkt.id), tkt['summary']))
                if other_tkts:
                    msg = tag.p("Other %s tickets (other versions) are still open on the same url and must be closed first: " % ticket['type'],
                                tag.p([tag.a("#%s (%s) " % (tid, summary),
                                             href=req.href('ticket', tid))
                                       for (tid, summary) in other_tkts]))
                    issues.append((None, msg))

            # Checks previous version closed before creating a new one
            if ((ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy and
                 ticket['ecmtype'] == 'Technical Note' and
                 ticket['fromecm'] != 'New Technical Note') or
                (ticket['type'] == 'FEE' and
                 ticket['fromfee'] != 'New Evolution Sheet')):
                ci_name = ticket['summary'].rsplit('_v')[0]

                # Check sourceurl not the same as other open tickets on same CI
                db = self.env.get_db_cnx()
                cursor = db.cursor()
                cursor.execute("SELECT id FROM ticket t,ticket_custom tc "
                               "WHERE t.type='%s' "
                               "AND t.status <> 'closed' "
                               "AND t.id=tc.ticket "
                               "AND tc.name='configurationitem' "
                               "AND tc.value='%s'" % (ticket['type'], ci_name))
                other_tkts = []
                for row in cursor:
                    tkt = Ticket(self.env, row[0])
                    other_tkts.append((str(tkt.id), tkt['summary']))
                if other_tkts:
                    msg = tag.p("Other %s tickets (other versions) are still open and must be closed first: " % ticket['type'],
                                tag.p([tag.a("#%s (%s) " % (tid, summary),
                                             href=req.href('ticket', tid))
                                       for (tid, summary) in other_tkts]))
                    issues.append((None, msg))

            # Checks previous version closed before creating a new one
            if (ticket['type'] == 'FEE' and
                ticket['fromfee'] != 'New Evolution Sheet'):
                ci_name = ticket['summary'].rsplit('_v')[0]

                # Check sourceurl not the same as other open tickets on same CI
                db = self.env.get_db_cnx()
                cursor = db.cursor()
                cursor.execute("SELECT id FROM ticket t,ticket_custom tc "
                               "WHERE t.type='FEE' "
                               "AND t.status <> 'closed' "
                               "AND t.id=tc.ticket "
                               "AND tc.name='configurationitem' "
                               "AND tc.value='%s'" % ci_name)
                other_tkts = []
                for row in cursor:
                    tkt = Ticket(self.env, row[0])
                    other_tkts.append((str(tkt.id), tkt['summary']))
                if other_tkts:
                    msg = tag.p("Other %s tickets (other versions) are still open and must be closed first: " % ticket['type'],
                                tag.p([tag.a("#%s (%s) " % (tid, summary),
                                             href=req.href('ticket', tid))
                                       for (tid, summary) in other_tkts]))
                    issues.append((None, msg))

        # -------------
        # Ticket exists
        # -------------

        else:
            if 'status' in ticket._old and ticket._old['status'] == 'closed':
                ticket['resolution'] = None

            if ticket['type'] in ('EFR', 'ECR', 'RF', 'PRF'):
                # Ticket closure decision ground
                if 'comment' in req.args:
                    comment = req.args.get('comment')
                else:
                    comment = ticket['submitcomment']
                if (ticket['status'] == 'closed' and
                    not form.TicketForm.get_closure_decision_ground(
                        self.env, ticket, comment)):
                    if (ticket['type'] in ('RF', 'PRF') and
                            ticket['resolution'] == 'fixed'):
                        # For RF/PRF closure as fixed, the associated changeset
                        # is known by TRAC so an appropriate comment can be forged
                        # in case the user didn't supply one
                        rev = ''
                        version_search = ticket['description'].rfind('[[BR]]Document')
                        if version_search != -1:
                            documentlink = ticket['description'][version_search:]
                            m = re.search('\[/browser(?:/(\w+))?/(?:trunk|tags|'
                                          'branches)/.+? (\d+)\]', documentlink)
                            if m:
                                reponame = m.group(1)
                                rev = m.group(2)

                                # Update submit comment
                                modified_comment = 'changeset:%s' % rev
                                if reponame:
                                    modified_comment += '/%s' % reponame
                                if comment:
                                    modified_comment += ' %s' % comment
                                ticket['submitcomment'] = modified_comment
                                req.args['comment'] = modified_comment
                    elif ('status' not in ticket._old and
                          'resolution' not in ticket._old):
                        # if ticket already closed
                        # only when changing the resolution
                        # shall an appropriate comment be supplied
                        pass
                    else:
                        label = web_ui.Ticket_UI.field_label(ticket, 'submitcomment')
                        msg = ('When closing a ticket or modifying the resolution'
                               ' of a closed ticket, an appropriate comment'
                               ' is required. Point your mouse pointer on'
                               ' the question mark to see what comment is expected'
                               ' and re-submit.')
                        issues.append((label, msg))

            # Checks on PRF remarks sheet before transitioning
            if (ticket['type'] == 'PRF' and 'status' in ticket._old):
                label = web_ui.Ticket_UI.field_label(ticket, 'sourcefile')
                tp_data = form.TicketForm.get_ticket_process_data(self.env, req.authname, ticket)
                tf = tp_data['ticket_form']
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    import openpyxl
                    wbObj = openpyxl.load_workbook(tf.content_filename)
                    # Remarks sheet
                    wbk = wbObj.worksheets[1]
                    if (ticket._old['status'] == '01-assigned_for_description' and
                        ticket['status'] == '07-assigned_for_closure_actions'):
                        # Missing "No remark"
                        if wbk.cell(row=3, column=3).value != _("No remark"):
                            msg = _('When a document is accepted as is, '
                                    'the sentence "No remark" is expected as first remark. '
                                    'This indicate that the review has been completed.')
                            issues.append((label, msg))
                    elif (ticket._old['status'] == '01-assigned_for_description' and
                          ticket['status'] == '03-assigned_for_analysis'):
                        for rowNum in range(3, wbk.max_row + 1):
                            # Only lines with remarks
                            if wbk.cell(row=rowNum, column=3).value is not None:
                                # Missing chapter and/or page
                                if wbk.cell(row=rowNum, column=2).value is None:
                                    msg = _('To make it easier to take your comments into account, '
                                            'please enter the page or paragraph concerned by your comment #%(row)s.', row=(rowNum-2))
                                    issues.append((label, msg))
                                # Missing criticality
                                if wbk.cell(row=rowNum, column=4).value is None:
                                    msg = _('To emphasize the importance of your comment #%(row)s, '
                                            'please enter the criticality of the point raised.', row=(rowNum-2))
                                    issues.append((label, msg))
                    elif (ticket._old['status'] == '03-assigned_for_analysis' and
                          ticket['status'] == '04-analysed'):
                        for rowNum in range(3, wbk.max_row + 1):
                            # Only lines with remarks
                            if wbk.cell(row=rowNum, column=3).value is not None:
                                # Missing decision
                                if wbk.cell(row=rowNum, column=5).value is None:
                                    msg = _('You must indicate your decision whether to take the remark #%(row)s into account. '
                                            'If OK, the document will be modified, '
                                            'else (NOK) do explain why you reject the implementation. '
                                            'Implementation may be rejected even if the remark is relevant.', row=(rowNum-2))
                                    issues.append((label, msg))
                                # Missing reject explanation
                                elif (wbk.cell(row=rowNum, column=5).value == 'NOK' and
                                    wbk.cell(row=rowNum, column=6).value is None):
                                    msg = _("As the remark #%(row)s is NOK'ed (will not be implemented), "
                                            'an explanation why it is rejected is required.', row=(rowNum-2))
                                    issues.append((label, msg))
                    elif (ticket._old['status'] == '04-analysed' and
                          ticket['status'] == '07-assigned_for_closure_actions'):
                        for rowNum in range(3, wbk.max_row + 1):                                    
                            # Only lines with remarks
                            if wbk.cell(row=rowNum, column=3).value is not None:
                                # Missing implementation
                                if wbk.cell(row=rowNum, column=5).value == 'OK':
                                    msg = _('You cannot accept the document as is because the document is to be modified '
                                            "to implement OK'ed remark #%(row)s ", row=(rowNum-2))
                                    issues.append((label, msg))
                    elif (ticket._old['status'] == '06-implemented' and
                          ticket['status'] == '07-assigned_for_closure_actions'):
                        for rowNum in range(3, wbk.max_row + 1):                                    
                            # Only lines with remarks
                            if wbk.cell(row=rowNum, column=3).value is not None:
                                # Missing revision number
                                if (wbk.cell(row=rowNum, column=5).value == 'OK' and
                                    wbk.cell(row=rowNum, column=7).value is None):
                                    msg = _('You have to indicate the revision number where you verified '
                                            "the correct implementation of OK'ed remark #%(row)s ", row=(rowNum-2))
                                    issues.append((label, msg))
                                elif (wbk.cell(row=rowNum, column=5).value == 'NOK' and
                                    wbk.cell(row=rowNum, column=7).value is not None):
                                    msg = _('No revision number is to be indicated for remark #%(row)s '
                                            "as it has not been OK'ed for implementation.", row=(rowNum-2))
                                    issues.append((label, msg))                                    

            # Check validity of parent MOM
            if ticket['type'] == 'DOC' and ticket['parent']:
                # Don't know if other plugin validity check
                # will be done before or after mine so call it now
                # it will be done twice...
                TracchildticketsModule(self.env).validate_ticket(req, ticket)
                # Check parent ticket is CCB MOM
                msg = None
                try:
                    tid = int(ticket.values.get('parent').lstrip('#'))
                    tkt = Ticket(self.env, tid)
                    if tkt['type'] != 'MOM':
                        msg = "Ticket %s is not a MOM" % tid
                    elif tkt['momtype'] != 'CCB':
                        msg = "Ticket %s is not a CCB MOM" % tid
                    elif tkt['skill'] not in util.get_milestone_skills(self.env, ticket['skill']):
                        msg = "Ticket %s is not a CCB MOM for skill %s" % (
                            tid,
                            ticket['skill'])
                    else:
                        # The  milestone tag shall include the DOC ticket
                        milestonetag = model.Tag(self.env, name=tkt['milestonetag'])
                        for baseline_item in model.BaselineItem.select(
                                 self.env, ['baselined_tag="' + milestonetag.name + '"']):
                            tg = model.Tag(self.env, name=baseline_item.name)
                            if util.get_doc_tktid(self.env, tg.tagged_item) == ticket.id:
                                break
                        else:
                            msg = "Ticket %s is not included in referenced milestone tag of ticket %s" % (
                                ticket.id, tid)

                except Exception:
                    e = sys.exc_info()[1]
                    try:
                        msg = exception_to_unicode(e)
                    finally:
                        del e
                if msg:
                    label = web_ui.Ticket_UI.field_label(ticket, 'parent')
                    issues.append((label, msg))

            # Checks on PDF before leaving edition
            if ((ticket['type'] == 'DOC' or (ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy) or ticket['type'] == 'FEE') and
                'status' in ticket._old and ticket._old['status'] == '01-assigned_for_edition' and
                not (ticket['status'] == 'closed' and ticket['resolution'] == 'rejected')):
                if ticket['pdffile'] != 'N/A':
                    if not ticket['pdffile']:
                        label = web_ui.Ticket_UI.field_label(ticket, 'pdffile')
                        msg = "No PDF file was selected - maybe none is available !"
                        issues.append((label, msg))
                    else:
                        # So a signature is expected
                        if ticket['sourcefile'] and ticket['sourcefile'] != 'N/A':
                            # Check PDF file is newer than source file
                            source_url = util.get_url(ticket['sourceurl'])
                            sourcefile_url = '%s/%s' % (source_url, ticket['sourcefile'])
                            pdffile_url = '%s/%s' % (source_url, ticket['pdffile'])
                            sourcefile_rev = util.get_last_path_rev_author(
                                self.env, sourcefile_url)[2]
                            pdffile_rev = util.get_last_path_rev_author(
                                self.env, pdffile_url)[2]
                            if int(pdffile_rev) < int(sourcefile_rev):
                                msg = tag.p("It seems the PDF file is older than the source file ",
                                            tag.b("in the repository"),
                                            " - have you (re)generated ",
                                            "and/or submitted the PDF file ",
                                            "before leaving the edition state ?")
                                issues.append((None, msg))
                            # Check PDF file is not empty
                            repos = util.get_repository(self.env, pdffile_url)
                            node = repos.get_node(pdffile_url, pdffile_rev)
                            if node.get_content_length() == 0:
                                msg = tag.p("It seems the PDF file is empty ",
                                            tag.b("in the repository"),
                                            " - have you (re)generated ",
                                            "and/or submitted the PDF file ",
                                            "before leaving the edition state ?")
                                issues.append((None, msg))

            # Check coherence between authors list and ticket owner when signing as author
            if ((ticket['type'] == 'DOC' or
                 (ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy) or
                 ticket['type'] == 'FEE') and
                'status' in ticket._old and ticket._old['status'] == '01-assigned_for_edition' and
                action != 'return_to_peer_review' and action != 'abort_ticket'):
                owner = ticket['owner'] if 'owner' not in ticket._old else ticket._old['owner']
                authors = TicketWF.get_WF(ticket).get_authors(ticket)
                if owner not in authors:
                    msg = tag.p("You may not sign as author ",
                                "as you are not in the authors list. ",
                                "It seems you have not modified the %s " % ticket['type'],
                                "by doing a lock/unlock cycle ",
                                "and/or submitted new data.")
                    issues.append((None, msg))

            # Prevent an admin Trac to sign unless he/she has a role in the project
            if ((ticket['type'] == 'DOC' or
                 (ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy) or
                 ticket['type'] == 'FEE') and
                 (('status' in ticket._old and ticket._old['status'] == '01-assigned_for_edition') or
                  action in ('approve', 'optional_approve', 'approve_fee', 'send', 'send_fee')) and
                 action != 'return_to_peer_review'): # DOC
                owner = ticket['owner'] if 'owner' not in ticket._old else ticket._old['owner']
                users_permissions = util.Users(self.env)
                if (owner in users_permissions.users_by_profile['admin'] and
                    owner not in chain.from_iterable(users_permissions.users_by_role.values())):
                    msg = tag.p("You may not sign as a Trac admin, ",
                                "a role in the project is required.")
                    issues.append((None, msg))

            # Check coherence between source header and transition
            if (ticket['type'] == 'DOC' and
                action in ('assign_for_peer_review', 'assign_for_formal_review') and
                ticket['sourcefile'].endswith('.docm')):
                template_cls = cache.Ticket_Cache.get_subclass(ticket['type'])
                with template_cls(self.env,
                                  self.trac_env_name,
                                  req.authname,
                                  ticket) as doc:
                    doc.checkout(ticket['sourcefile'])
                    version_status = doc.get_version_status(ticket['sourcefile'])
                    if action == 'assign_for_peer_review':
                        if not version_status.startswith('Draft'):
                            label = web_ui.Ticket_UI.field_label(ticket, 'pdffile')
                            msg = "A Draft status is required for the document"
                            issues.append((label, msg))
                    else:
                        if version_status != 'Released':
                            label = web_ui.Ticket_UI.field_label(ticket, 'pdffile')
                            msg = "A Released status is required for the document"
                            issues.append((label, msg))

            # Check coherence between parent ECR status and transition
            if (ticket['type'] == 'DOC' and action == 'assign_for_formal_review' and
                ticket['blocking']):
                tids = [tid.strip()
                        for tid in ticket['blocking'].split(',')]
                for tid in tids:
                    msg = None
                    try:
                        tkt = Ticket(self.env, tid)
                        if tkt['status'] < '05-assigned_for_implementation':
                            msg = "Ticket %s should have reached 05-assigned_for_implementation status" % tid
                    except Exception:
                        e = sys.exc_info()[1]
                        try:
                            msg = exception_to_unicode(e)
                        finally:
                            del e
                    if msg:
                        label = web_ui.Ticket_UI.field_label(ticket, 'blocking')
                        issues.append((label, msg))

            # Check required fields are not empty
            if (ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy) or ticket['type'] == 'FEE':
                if action in ('send', 'send_fee'):
                    for field in web_ui.Ticket_UI.get_UI(ticket)(self, ticket).required_fields():
                        if not ticket[field]:
                            label = web_ui.Ticket_UI.field_label(ticket, field)
                            msg = "this field is required"
                            issues.append((label, msg))

            if ((ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy and action == 'send') or
                (ticket['type'] == 'FEE' and action == 'send_fee')):
                # Check email addresses are valid
                for field in ['fromemail', 'toemail']:
                    if ticket[field]:
                        email = ticket[field]
                        try:
                            # validate and get info
                            v = validate_email(email, allow_smtputf8=False)
                            # replace with normalized form
                            ticket[field] = v["email"]
                        except EmailNotValidError as e:
                            # email is not valid, exception message is human-readable
                            label = web_ui.Ticket_UI.field_label(ticket, field)
                            msg = "validation of email address failed: %s" % str(e)
                            issues.append((label, msg))

                for field in ['carboncopy']:
                    if ticket[field]:
                        label = web_ui.Ticket_UI.field_label(ticket, field)
                        regexp = r"\A\s*([^(]+?)\s*\(\s*([^)]+?)\s*\)\s*\Z"
                        lines = []
                        issue = False
                        for line in ticket['carboncopy'].splitlines():
                            match = re.search(regexp, line)
                            if not match:
                                msg = "Incorrect syntax (%s)" % regexp
                                issues.append((label, msg))
                                issue = True
                            else:
                                name = match.group(1)
                                if not name:
                                    msg = "Missing name (%s)" % line
                                    issues.append((label, msg))
                                    issue = True
                                email = match.group(2)
                                if not email:
                                    msg = "Missing email address (%s)" % line
                                    issues.append((label, msg))
                                    issue = True
                                try:
                                    # validate and get info
                                    v = validate_email(email, allow_smtputf8=False)
                                    # replace with normalized form
                                    lines.append("%s (%s)" % (name, v["email"]))
                                except EmailNotValidError as e:
                                    # email is not valid, exception message is human-readable
                                    label = web_ui.Ticket_UI.field_label(ticket, field)
                                    msg = "validation of email address failed: %s" % str(e)
                                    issues.append((label, msg))
                                    issue = True
                        if not issue:
                            ticket[field] = '\n'.join(lines)

            # Check child PRF are closed
            if ((ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy and action in
                 ('assign_for_optional_review', 'assign_for_optional_approval', 'assign_for_sending', 'abort_ticket')) or
                (ticket['type'] == 'FEE' and action in
                 ('assign_for_fee_review_management', 'assign_for_fee_internal_approval_management', 'abort_ticket')) or
                (ticket['type'] == 'DOC' and action in
                 ('assign_for_peer_review', 'assign_for_formal_review', 'assign_for_approval', 'abort_ticket'))):
                # Child tickets for current version status
                childtickets = util.child_tickets_for_tag(ticket)
                # All child PRF tickets closed for the current version status
                for tkt in childtickets:
                    if tkt['status'] != 'closed':
                        msg = tag.p("Child ticket ",
                                    tag.a("#%s" % tkt.id,
                                          href=req.href('ticket', tkt.id)),
                                    " (",
                                    tag.a("%s" % tkt['summary'],
                                          href=req.href('ticket', tkt.id)),
                                    ") shall be closed first !")
                        issues.append((None, msg))

            # Check there is at least one child PRF ticket with resolution fixed
            if ticket['type'] == 'DOC' and action == 'assign_for_approval':
                # Child tickets for current version status
                childtickets = util.child_tickets_for_tag(ticket)
                # At least one child PRF ticket for the current Proposed version
                if not childtickets:
                    msg = tag.p("A child PRF ticket on current Proposed status "
                                "is required for approval")
                    issues.append((None, msg))
                # At least one child PRF ticket with resolution fixed
                elif not next((tkt.id for tkt in childtickets
                               if tkt['resolution'] == 'fixed'), []):
                    msg = tag.p("A child PRF ticket on current Proposed status "
                                "with resolution fixed is required for approval")
                    issues.append((None, msg))

            # Check parent MOM
            if ticket['type'] == 'DOC' and action == 'release':
                # A parent MOM is required if CC1/HC1
                if (not ticket['parent'] and
                    ticket['controlcategory'] == 'CC1/HC1'):
                    label = web_ui.Ticket_UI.field_label(ticket, 'parent')
                    msg = "A parent MOM is required when CC1/HC1 is selected"
                    issues.append((label, msg))

            # Check if version tag can be removed
            try:
                tg = model.Tag(self.env, ticket['document']) if ticket['document'] else None
                if (tg and
                    ((ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy and
                      action == 'abort_sending') or
                     (ticket['type'] == 'FEE' and
                      action == 'abort_fee_customer_approval_management') or
                     (ticket['type'] == 'DOC' and
                      ((action == 'abort_peer_review' and tg.status == 'Draft') or
                       (action == 'abort_formal_review' and tg.status == 'Proposed') or
                       (action == 'reopen' and ticket._old['resolution'] != 'rejected' and
                        tg.status == 'Released'))))):
                        db = self.env.get_db_cnx()
                        # The version tag can only be removed
                        # if it is not included in a baseline
                        blocking = [(v.name, v.baselined_tag)
                                    for v in model.BaselineItem.select(
                                    self.env,
                                    ['name="' + tg.name + '"'], db=db)]
                        msg = None
                        if blocking:
                            msg = tag()
                            for block in blocking:
                                w = model.Tag(self.env, block[1], db=db)
                                if w.tag_url is None:
                                    tagged = '- not yet applied - '
                                else:
                                    tagged = '- applied - '
                                if w.review is None:
                                    including_page = admin.web_ui.VersionTagsAdminPanel.page_type
                                else:
                                    including_page = admin.web_ui.MilestoneTagsAdminPanel.page_type
                                msg(tag.p("Can't remove ",
                                          tag.em("included tag "),
                                          tag.a("%s" % block[0],
                                                href=req.href.admin('tags_mgmt',
                                                                    'version_tags',
                                                                    block[0])),
                                          " because it is included in ",
                                          tagged,
                                          tag.em("baselined tag "),
                                          tag.a("%s." % block[1],
                                                href=req.href.admin('tags_mgmt',
                                                                    including_page,
                                                                    block[1],
                                                                    selected_item="%s" % block[0]))))
                            msg(tag.p("Either remove the listed baselined tags"
                                      " or remove the problematic included tags from them "
                                      "(if the baselined tags are not yet applied).",
                                      class_="message"))
                        if msg:
                            issues.append((None, msg))

                        # The version tag can only be removed
                        # if it is not used as baseline in one or more ticket(s)
                        # other than the current ticket !
                        block = None
                        cursor = db.cursor()
                        cursor.execute("SELECT ticket FROM ticket_custom "
                                       "WHERE name='document' "
                                       "and value='%s' and ticket<>%s" %
                                       (tg.name, ticket.id))
                        tkt_ids = [int(row[0]) for row in cursor]
                        if tkt_ids:
                            block = (tg.name, len(tkt_ids))
                        msg = None
                        if block:
                            msg = tag()
                            msg(tag.p("Can't remove ",
                                      tag.em("version tag "),
                                      tag.a("%s" % block[0],
                                            href=req.href.admin('tags_mgmt',
                                                                'version_tags',
                                                                block[0])),
                                      " because it is used as a baseline tag "
                                      "in ",
                                      tag.a("%d " % block[1],
                                            href=req.href.query(
                                                group="status",
                                                document="%s" % block[0],
                                                id="!%s" % ticket.id,
                                                order="priority",
                                                col=["id",
                                                     "summary",
                                                     "type",
                                                     "document"])),
                                      "ticket(s)."))
                            msg(tag.p("If feasible, change the baseline tag(s) "
                                      "for the ticket(s) involved.",
                                      class_="message"))
                        if msg:
                            issues.append((None, msg))

                        # The version tag can only be removed
                        # if it is not used as baseline for branching
                        blocking = [(branch.source_tag, branch.id)
                                    for branch in model.Branch.select(
                                    self.env,
                                    ['source_tag="' + tg.name + '"'],
                                    db=db)]
                        msg = None
                        if blocking:
                            msg = tag()
                            for block in blocking:
                                w = model.Branch(self.env, block[1], db=db)
                                if w.branch_url is None:
                                    tagged = '- not yet applied - '
                                else:
                                    tagged = '- applied - '
                                msg(tag.p("Can't remove ",
                                          tag.em("version tag "),
                                          tag.a("%s" % block[0],
                                                href=req.href.admin(
                                                'tags_mgmt', 'version_tags', block[0])),
                                          " because it is the source tag of ",
                                          tagged,
                                          tag.em("branch "),
                                          tag.a("B%s." % block[1],
                                                href=req.href.admin(
                                                'tags_mgmt', 'branches', block[1]))))
                            msg(tag.p("If feasible, remove the listed branches "
                                      "or change the source tag of them "
                                      "(if the branches are not yet applied).",
                                      class_="message"))
                        if msg:
                            issues.append((None, msg))
            except ResourceNotFound:
                pass

            # Check source file is not locked so commit will be possible
            if (((ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy and action in ['reassign', 'reassign_for_edition']) or
                 (ticket['type'] == 'FEE' and action in ['reassign', 'reassign_for_fee_edition']) or
                 (ticket['type'] == 'DOC' and action in ['reassign', 'assign_for_edition'])) and
                ticket['sourcefile'] and ticket['sourcefile'].endswith('.docm')):

                template_cls = cache.Ticket_Cache.get_subclass(ticket['type'])
                with template_cls(self.env,
                                  self.trac_env_name,
                                  req.authname,
                                  ticket) as doc:
                    msg = None
                    label = web_ui.Ticket_UI.field_label(ticket, 'sourcefile')
                    repos_status = doc.status(ticket['sourcefile'], 'repos-status')
                    wc_status = doc.status(ticket['sourcefile'], 'wc-status')
                    if repos_status['lock_agent']:
                        if repos_status['lock_agent'] == 'trac':
                            if wc_status['lock_agent'] is None:
                                msg = "The file %s is locked by %s with ticket %s. Remove that lock - with the ticket Unlock function - so that Trac can proceed" % (
                                    ticket['sourcefile'], repos_status['lock_client'], repos_status['lock_ticket'])
                        else:
                            msg = "The file %s is locked outside of trac by %s. Remove that lock - eg with TortoiseSVN Release lock function - so that Trac can proceed" % (
                                ticket['sourcefile'], repos_status['lock_agent'])
                    if msg:
                        issues.append((label, msg))

            # Check PDF file is not locked so commit will be possible after signing
            if ((ticket['type'] == 'DOC' or (ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy) or ticket['type'] == 'FEE') and
                ticket['pdffile'] and ticket['pdffile'] != 'N/A'):
                workflow = self.action_controllers[0]
                ticket_wf = workflow.get_ticket_wf(ticket)(workflow, req, ticket, action)
                previous_status = 'status' in ticket._old and ticket._old['status'] or ticket['status']
                # legacy support
                if previous_status in ("04-assigned_for_release", "05-assigned_for_release"):
                    previous_status = "06-assigned_for_release"
                if ticket_wf.signing_in_operations(previous_status):
                    template_cls = cache.Ticket_Cache.get_subclass(ticket['type'])
                    with template_cls(self.env,
                                      self.trac_env_name,
                                      req.authname,
                                      ticket) as doc:
                        msg = None
                        label = web_ui.Ticket_UI.field_label(ticket, 'pdffile')
                        repos_status = doc.status(ticket['pdffile'], 'repos-status')
                        wc_status = doc.status(ticket['pdffile'], 'wc-status')
                        if repos_status['lock_agent']:
                            if repos_status['lock_agent'] == 'trac':
                                if wc_status['lock_agent'] is None:
                                    msg = "The file %s is locked by %s with ticket %s. Remove that lock - with the ticket Unlock function - so that Trac can proceed" % (
                                        ticket['pdffile'], repos_status['lock_client'], repos_status['lock_ticket'])
                            else:
                                msg = "The file %s is locked outside of trac by %s. Remove that lock - eg with TortoiseSVN Release lock function- so that Trac can proceed" % (
                                    ticket['pdffile'], repos_status['lock_agent'])
                        if msg:
                            issues.append((label, msg))

            # Check the archives have not been removed
            if (ticket["type"] == "ECM" and not web_ui.Ticket_UI.get_UI(ticket).legacy and
                ticket["ecmtype"] == "Document Delivery" and "status" in ticket._old and action != 'abort_ticket'):
                if cache.PDFPackage.get_archives_number(ticket) == 0:
                    msg = tag.p("No archives were found for current ticket %s." % ticket["summary"])
                    if ticket._old["status"] == "01-assigned_for_edition":
                        msg(tag.p("Use the <Select documents> button in the ticket header "
                                  "to select the documents."))
                    issues.append((None, msg))
            if (ticket['type'] == 'MOM' and
                ticket['momtype'] in ('CCB', 'Review') and
                action == 'resolve' and
                ticket['resolution'] != 'rejected'):
                # Check associated milestonetag is applied
                tg = model.Tag(self.env, ticket['milestonetag'])
                if not tg.tag_url:
                    msg = tag.span('The associated milestonetag ', tag.a(ticket['milestonetag'], href=req.href.admin('tags_mgmt', 'milestone_tags', ticket['milestonetag'])))
                    msg += tag.span(' has not been applied. It is mandatory for closing this MOM ticket as %s.' % ticket['resolution'])
                    issues.append((None, msg))

        # -----------------------------
        # Ticket in creation or created
        # -----------------------------

        # Check Parent ECR(s) existence if CC1/HC1
        if ticket['type'] == 'DOC' and ticket['controlcategory'] == 'CC1/HC1':
            # If there are released versions then it is mandatory
            status = [tg.status for tg in model.Tag.select(
                      self.env, ['tracked_item="%s"' % ticket['configurationitem'], 'status="Released"'],
                      tag_type='version_tags')]
            if 'Released' in status and not(len(status) == 1 and action == 'reopen'):
                # Check existence of parent ECR(s)
                if not ticket['blocking']:
                        label = web_ui.Ticket_UI.field_label(ticket, 'blocking')
                        msg = ("ECR(s) are required when the control category is CC1/HC1 "
                               "and the document has already been released")
                        issues.append((label, msg))

                # Check validity of parent ECR(s)
                else:
                    tids = [tid.strip()
                            for tid in ticket['blocking'].split(',')]
                    for tid in tids:
                        msg = None
                        try:
                            tkt = Ticket(self.env, tid)
                            if tkt['type'] != 'ECR':
                                msg = tag()
                                msg('Ticket ',
                                    tag.a("#%s" % tid,
                                          href=req.href.ticket(tid),
                                          title=tkt['summary']),
                                    ' is not an ECR (%s)' % tkt['type'])
                            elif tkt['status'] == 'closed':
                                msg = tag()
                                msg('Ticket ',
                                    tag.a("#%s" % tid,
                                          href=req.href.ticket(tid),
                                          title=tkt['summary']),
                                    ' is closed')
                        except Exception:
                            e = sys.exc_info()[1]
                            try:
                                msg = exception_to_unicode(e)
                            finally:
                                del e
                        if msg:
                            label = web_ui.Ticket_UI.field_label(ticket, 'blocking')
                            issues.append((label, msg))

        return issues

    # IMasterObserver method

    def master_validate(self, req, ticket):

        issues = []
        dc_url = self.env.config.get('artusplugin', 'dc_url')

        # Check child tickets type for EFR ticket
        if ticket['type'] == 'EFR':
            field = 'blockedby'
            field_label = web_ui.Ticket_UI.field_label(ticket, 'blockedby')
            # The blockedby field has already been validated
            # by the MasterTickets plugin and every remaining
            # ticket number is a valid one

            # Check every child is a DOC or an ECR
            tkt_ids = (MasterTicketsSystem.NUMBERS_RE).findall(ticket[field] or '')
            for tkt_id in tkt_ids[:]:
                tkt = Ticket(self.env, tkt_id)
                if tkt['type'] not in ('DOC', 'ECR'):
                    msg = tag()
                    msg('Ticket ',
                        tag.a("#%s" % tkt_id,
                              href=req.href.ticket(tkt_id),
                              title=tkt['summary']),
                        ' is not a DOC or an ECR (%s)' % tkt['type'])
                    issues.append((field_label, msg))
                if issues:
                    msg = tag()
                    msg('Every child ticket of an EFR '
                        'is expected to be a DOC or an ECR. See ',
                        tag.a("What is an EFR/ECR",
                              href="%s/index.php?post/74" % dc_url))
                    issues.append((None, msg))

        # Check parent tickets type for ECR ticket
        if ticket['type'] == 'ECR':
            field = 'blocking'
            # The blocking field has already been validated
            # by the MasterTickets plugin and every remaining
            # ticket number is a valid one
            if ticket['ecrtype'] == 'Evolution':
                # field label not defined elsewhere in python (it is set in js)
                field_label = 'Parent ECR(s)'
                # Check every parent is an ECR
                tkt_ids = (MasterTicketsSystem.NUMBERS_RE).findall(ticket[field] or '')
                for tkt_id in tkt_ids[:]:
                    tkt = Ticket(self.env, tkt_id)
                    if tkt['type'] != 'ECR':
                        msg = tag()
                        msg('Ticket ',
                            tag.a("#%s" % tkt_id,
                                  href=req.href.ticket(tkt_id),
                                  title=tkt['summary']),
                            ' is not an ECR (%s)' % tkt['type'])
                        issues.append((field_label, msg))
                if issues:
                    msg = tag()
                    msg('Every parent ticket of an ECR of type '
                        'Evolution is expected to be an ECR. See ',
                        tag.a("What is an EFR/ECR",
                              href="%s/index.php?post/74" % dc_url))
                    issues.append((None, msg))

            elif ticket['ecrtype'] == 'Problem Report':
                # field label not defined elsewhere in python (it is set in js)
                field_label = 'Parent EFR(s)'
                # Check there is at least one parent and every parent is an EFR
                tkt_ids = (MasterTicketsSystem.NUMBERS_RE).findall(ticket[field] or '')
                if tkt_ids:
                    for tkt_id in tkt_ids[:]:
                        tkt = Ticket(self.env, tkt_id)
                        if tkt['type'] != 'EFR':
                            msg = tag()
                            msg('Ticket ',
                                tag.a("#%s" % tkt_id,
                                      href=req.href.ticket(tkt_id),
                                      title=tkt['summary']),
                                ' is not an EFR (%s)' % tkt['type'])
                            issues.append((field_label, msg))
                    if issues:
                        msg = tag()
                        msg('Every parent ticket of an ECR of type '
                            'Problem Report is expected to be an EFR. See ',
                            tag.a("What is an EFR/ECR",
                                  href="%s/index.php?post/74" % dc_url))
                        issues.append((None, msg))
                else:
                    msg = tag()
                    msg('No parent EFR is specified. '
                        'A parent EFR at least is required '
                        'for an ECR of type Problem Report. See ',
                        tag.a("What is an EFR/ECR",
                              href="%s/index.php?post/74" % dc_url))
                    issues.append((None, msg))

        return issues

    # ITicketChangeListener methods

    def ticket_created(self, ticket):
        """Called when a ticket is created."""

        # authname is necessary for addressing the correct working copy
        req = util.get_req()
        authname = req.authname if req else 'trac'
        if authname == 'trac':
            return

        syslog.syslog("%s(%s): Creation ticket %s (%s)" % (self.trac_env_name, authname, ticket.id, ticket['type']))

        # Set ID for tickets WITH chronological number
        if (ticket['type'] not in ('RF', 'PRF', 'DOC') and
            (ticket['type'] != 'MOM' or
             (ticket['momtype'] not in ('CCB', 'Progress', 'Internal') and
              (ticket['momtype'] != 'Review' or
               ticket['skill'] not in self._skills['proj'])))):
            now = datetime.now(localtz)
            ticket_chrono_filename = self._get_ticket_chrono_filename(ticket, now)
            if ticket['type'] in ('EFR', 'ECR'):
                ticket_id = "%s_%s" % (ticket['type'], self.program_name)
                number = util.new_ticket_number(ticket_chrono_filename)

                if ticket['type'] == 'EFR':
                    skills = self.env.config.get('ticket-custom', 'skill.options')
                    if 'SYS' not in skills:
                        ticket_id = "%s_%s" % (ticket_id, ticket['skill'])
                else:
                    ticket_id = "%s_%s" % (ticket_id, ticket['skill'])

                summary = "%s_%s" % (ticket_id, "%0*d" % (3, number))

            elif ticket['type'] == 'MOM':
                if ticket['momtype'] == 'Review':
                    if ticket['skill'] not in self._skills['proj']:
                        number = util.new_ticket_number(ticket_chrono_filename)
                        # YY-ddd
                        summary = '%s_%s_%s_%s_%s' % (
                            ticket['type'],
                            self.program_name,
                            ticket['skill'],
                            'REV',
                            "%s-%0*d" % (now.strftime("%y"),
                                         3,
                                         number))
                elif ticket['momtype'] == 'Audit':
                    number = util.new_ticket_number(ticket_chrono_filename)
                    # YY-ddd
                    summary = '%s_%s_%s_%s_%s' % (
                        ticket['type'],
                        self.program_name,
                        ticket['skill'],
                        'AUD',
                        "%s-%0*d" % (now.strftime("%y"),
                                     3,
                                     number))

            elif ticket['type'] in ('RISK', 'AI'):
                number = util.new_ticket_number(ticket_chrono_filename)
                summary = '%s_%s_%s_%s' % (
                    ticket['type'],
                    self.program_name,
                    ticket['skill'],
                    "%0*d" % (4, number))

            elif ticket['type'] == 'MEMO':
                number = util.new_ticket_number(ticket_chrono_filename)
                summary = '%s_%s_%s_%s' % (
                    ticket['type'],
                    self.program_name,
                    ticket['skill'],
                    "%0*d" % (3, number))
            
            elif ticket['type'] == 'ECM':
                # New tickets are not legacy ECMs
                if ((ticket['ecmtype'] == 'Technical Note' and ticket['fromecm'] == 'New Technical Note') or
                    ticket['ecmtype'] == 'Document Delivery'):
                    number = util.new_ticket_number(ticket_chrono_filename)
                    ci_name = '%s_%s_%s' % (
                        ticket['type'],
                        self.program_name,
                        "%0*d" % (3, number))
                    if ticket['ecmtype'] == 'Technical Note':
                        summary = ci_name + ticket['versionsuffix']
                    else:
                        summary = ci_name
                else:
                    ci_name = ticket['summary'].rsplit('_v')[0]
                    summary = ticket['summary']
                repo_path = self.env.config.get('artusplugin', '%s_ecm_forms_dir' % ticket['skill'])
                sourceurl = '%s/%s' % (repo_path, ci_name)
            
            elif ticket['type'] == 'FEE':
                ci_name = ticket['summary'].rsplit('_v')[0]
                summary = ticket['summary']
                repo_path = self.env.config.get('artusplugin', 'fee_forms_dir')
                sourceurl = '%s/%s' % (repo_path, ci_name)

            with self.env.db_transaction as db:
                db("UPDATE ticket SET summary=%s WHERE id=%s",
                   (summary, ticket.id))
            ticket['summary'] = summary
            
            if ticket['type'] in ('ECM', 'FEE'):
                # New tickets are not legacy ECMs
                with self.env.db_transaction as db:
                    db("UPDATE ticket_custom SET value=%s WHERE ticket=%s AND name='configurationitem'",
                       (ci_name, ticket.id))
                    db("UPDATE ticket_custom SET value=%s WHERE ticket=%s AND name='sourceurl'",
                       (sourceurl, ticket.id))
                ticket['configurationitem'] = ci_name
                # sourceurl is required for cache creation (semaphore naming)
                ticket['sourceurl'] = sourceurl

        # Create document entry if first DOC ticket
        if ticket['type'] == 'DOC':
            try:
                # Do nothing if document already exists
                model.Document(self.env, ticket['configurationitem'])
            except ResourceNotFound:
                # Create document
                cc_options = self.env.config.get('ticket-custom',
                                                 'controlcategory.options')
                cc_values = [option.strip() for option in cc_options.split('|')]
                cc = cc_values.index(ticket['controlcategory'])
                sf_options = self.env.config.get('ticket-custom',
                                                 'submittedfor.options')
                sf_values = [option.strip() for option in sf_options.split('|')]
                sf = sf_values.index(ticket['submittedfor'])
                doc = model.Document(self.env)
                doc['name'] = ticket['configurationitem']
                doc['shortname'] = doc['name'].split('_')[-1]
                doc['controlcategory'] = cc
                doc['submittedfor'] = sf
                doc['independence'] = int(ticket['independence'])
                doc['sourcetype'] = ticket['sourcetype']
                doc['pdfsigned'] = int(ticket['pdfsigned'])
                doc.insert()

        # Setup of working copy
        if ticket['type'] in ('ECM', 'FEE', 'DOC'):
            # New tickets are not legacy ECMs
            template_cls = cache.Ticket_Cache.get_subclass(ticket['type'])
            with template_cls(self.env, self.trac_env_name,
                              authname, ticket) as doc:
                if ticket['type'] == 'ECM':
                    ticket['sourcefile'] = '%s.docm' % ticket['configurationitem']
                    if ((ticket['ecmtype'] == 'Technical Note' and ticket['fromecm'] == 'New Technical Note') or
                        ticket['ecmtype'] == 'Document Delivery'):
                        doc.add(ticket['sourcefile'])
                    else:
                        doc.checkout(ticket['sourcefile'])
                    doc.lock(ticket['sourcefile'])
                    doc.update_data(ticket['sourcefile'])
                    doc.upgrade_document(ticket['sourcefile'])
                    ticket['pdffile'] = '%s.pdf' % ticket['configurationitem']
                    if ((ticket['ecmtype'] == 'Technical Note' and ticket['fromecm'] == 'New Technical Note') or
                        ticket['ecmtype'] == 'Document Delivery'):
                        doc.add(ticket['pdffile'])
                    else:
                        doc.checkout(ticket['pdffile'])
                    doc.lock(ticket['pdffile'])
                elif ticket['type'] == 'FEE':
                    ticket['sourcefile'] = '%s.docm' % ticket['configurationitem']
                    if ticket['fromfee'] == 'New Evolution Sheet':
                        doc.add(ticket['sourcefile'])
                    else:
                        doc.checkout(ticket['sourcefile'])
                    doc.lock(ticket['sourcefile'])
                    doc.update_data(ticket['sourcefile'])
                    doc.upgrade_document(ticket['sourcefile'])
                    ticket['pdffile'] = '%s.pdf' % ticket['configurationitem']
                    if ticket['fromfee'] == 'New Evolution Sheet':
                        doc.add(ticket['pdffile'])
                    else:
                        doc.checkout(ticket['pdffile'])
                    doc.lock(ticket['pdffile'])
                
                revision = doc.commit()
                # Existing ECM/FEE
                if revision == '' and ticket['type'] in ('ECM', 'FEE'):
                    # Remove the locks
                    doc.unlock(ticket['sourcefile'])
                    doc.unlock(ticket['pdffile'])
                    # Get revision
                    revision = util.get_last_path_rev_author(self.env, util.get_url(ticket['sourceurl']))[2]
                if revision:
                    ticket['sourceurl'] = '%s?rev=%s' % (
                        util.get_url(ticket['sourceurl']),
                        revision)
                    now = datetime.now(utc)
                    ticket.save_changes('trac', 'Ticket created', now)

        tickets_with_forms = set(('EFR', 'ECR', 'RF', 'PRF'))

        if 'MOM' in util.get_prop_values(self.env,
                                         'ticket_edit.office_suite').keys():
            tickets_with_forms.add('MOM')

        if ticket['type'] in tickets_with_forms:
            # Setup of temporary pseudo-working copy
            tp_data = form.TicketForm.get_ticket_process_data(self.env, authname, ticket)
            tft = tp_data['ticket_form_template']
            tf = tp_data['ticket_form']

            try:
                # Create pseudo-working copy root folder
                if os.access(tf.path, os.F_OK):
                    shutil.rmtree(tf.path)
                os.makedirs(tf.path)

                # Add webdav folder for remote edit
                if tf.office_suite == 'MS Office':
                    if not os.access(tf.content_path, os.F_OK):
                        os.mkdir(tf.content_path)

                # Add attachments folder
                if not os.access(tp_data['attachment'].destpath, os.F_OK):
                    os.makedirs(tp_data['attachment'].destpath)

                # Get form template
                if not os.access(tft.cache_path, os.F_OK):
                    os.makedirs(tft.cache_path)
                if tft.repo_form:
                    repos = util.get_repository(self.env, tft.source_path)
                    node = repos.get_node(tft.source_name, tft.source_rev)
                    templatefile = os.fdopen(os.open(tft.cache_name,
                                                     os.O_CREAT + os.O_WRONLY + os.O_TRUNC,
                                                     666), 'w')
                    shutil.copyfileobj(node.get_content(), templatefile)
                    templatefile.close()
                else:
                    shutil.copy(tft.source_name, tft.cache_name)

                # Create the empty form from the template form
                shutil.copy(tft.cache_name, tf.oldcontent_filename)

                # Copy the empty form into the edit form
                shutil.copy(tf.oldcontent_filename, tf.content_filename)

                if ticket['type'] == 'MOM':
                    # prepare the MOM form
                    skills = util.get_ticket_skills(self.env, ticket['skill'])
                    tf.setup(skills, ticket['milestonetag'])
                else:
                    # Update the edited form with TRAC and Office data
                    tf.update(ticket, tp_data, ['fields', None], None)

                # Copy the edited form into the empty form
                shutil.copy(tf.content_filename, tf.oldcontent_filename)

                # Define ignore property on 'ticket_oldcontent_path'
                with open('%s/.ignore' % tf.path, 'w') as ignorefile:
                    ignorefile.write(".trac\n*.edit\n*.mine\nattachments")
                    if tf.office_suite == 'MS Office':
                        ignorefile.write("\n*.xml\nwebdav")

                # Define ignore property on 'ticket_attachment_destpath'
                with open('%s/.ignore' % tp_data['attachment'].destpath, 'w') as ignorefile:
                    ignorefile.write("*.mine\n.bkup")

                # Creation of ticket record in the repository
                paths, dirs = util.analyse_url(self.env, tf.ticket_path)

                unix_cmd_list = [tp_data['svnmucc_template_cmd'] +
                                 '-m "%s" ' % _('ticket:%(id)s (on behalf of %(user)s)', id=tp_data['id'], user=authname) +
                                 ''.join(['mkdir "%s" ' % d for d in dirs]) +
                                 'mkdir "%s" ' % paths[-1] +
                                 'put "%s" "%s" ' % (tf.oldcontent_filename,
                                                     util.get_repo_url(self.env, tf.ticket_filename)) +
                                 'propsetf svn:ignore "%s/.ignore" "%s" ' % (
                                     tf.path, tf.http_url) +
                                 'mkdir "%s" ' % tp_data['attachment'].http_url +
                                 'propsetf svn:ignore "%s/.ignore" "%s" ' % (
                                     tp_data['attachment'].destpath, tp_data['attachment'].http_url)
                                 ]

                # Effective application of the list of commands
                retcode, lines = util.unix_cmd_apply(self.env, unix_cmd_list,
                                                     util.lineno())

                revision = ''
                if retcode == 0:
                    regular_expression = u'\A\s*r(\d+) committed [\S\s]+\Z'
                    match = re.search(regular_expression, lines[0])
                    if match:
                        revision = match.group(1)
                    # Update <TRAC to SVN> links
                    self._links_update(ticket,
                                       tp_data,
                                       revision,
                                       'creation',
                                       'Ticket created')
                else:
                    raise TracError("[Creation of ticket record in the repository]")

                # Removal of temporary pseudo-working copy
                shutil.rmtree(tf.path)

            except TracError:
                extraInfo = sys.exc_info()[1]
                try:
                    message = tag.p("Creation of %s " % tp_data['ticket_type'],
                                    tag.em("%s" % tp_data['ticket_id']),
                                    " has failed.",
                                    class_="message")
                    message(tag.p(extraInfo))
                    for line in lines:
                        message(tag.p(line))
                    raise TracError(message)
                finally:
                    del extraInfo
        return

    def ticket_changed(self, ticket, comment, author, old_values):
        """Called when a ticket is modified.

        `old_values` is a dictionary containing the previous values of the
        fields that have changed.
        """
        # authname is necessary for addressing the correct working copy
        if author:
            authname = author
        else:
            req = util.get_req()
            authname = req.authname if req else 'trac'
        if authname == 'trac':
            return

        tickets_with_forms = set(('EFR', 'ECR', 'RF', 'PRF'))

        if 'MOM' in util.get_prop_values(self.env,
                                         'ticket_edit.office_suite').keys():
            tickets_with_forms.add('MOM')

        if ticket['type'] in tickets_with_forms:
            tp_data = form.TicketForm.get_ticket_process_data(self.env, authname, ticket)
            tf = tp_data['ticket_form']
            if not util.exist_in_repo(self.env, tf.http_url):
                tickets_with_forms.remove(ticket['type'])

        if ticket['type'] in tickets_with_forms:

            # To avoid recursion
            if 'description' in old_values:
                return

            # any modification of trac data ?
            ticket_fields = tf.TICKET_FIELDS
            if ticket_fields:
                for field in old_values:
                    if field in ticket_fields.keys():
                        trac_data_modified = True
                        break
                else:
                    trac_data_modified = False
            else:
                trac_data_modified = False

            # Any modification of the edited form ?
            if os.access(tf.copycontent_path, os.F_OK):
                if filecmp.cmp(tf.content_filename, tf.copycontent_filename):
                    edit_form_modified = False
                else:
                    edit_form_modified = True
            else:
                edit_form_modified = False

            if edit_form_modified or trac_data_modified:
                #
                # Pre-processing associated with
                # form data or TRAC data modification
                #

                # Update of the working copy -
                # attachments may have been added, removed or
                # edited by someone else (Force Edit mode)
                unix_cmd_list = [tp_data['svn_template_cmd'] % {
                                 'subcommand': 'up'} +
                                 '"' +
                                 tf.path +
                                 '"']

                # Effective application of the list of commands
                retcode, lines = util.unix_cmd_apply(self.env, unix_cmd_list,
                                                     util.lineno())

                if retcode != 0:
                    raise TracError("[Update of the ticket form working copy]")

                # Update the form with TRAC and form data
                if ticket.time_created > tf.get_update_deployment_date():
                    tf.update(ticket,
                              tp_data,
                              ['fields', 'help_off'],
                              None,
                              old_values)

                # Copy the updated form into the form to check-in
                if os.access(tf.oldcontent_filename, os.F_OK):
                    os.remove(tf.oldcontent_filename)
                shutil.copy(tf.content_filename, tf.oldcontent_filename)

                #
                # Commit of the ticket form
                #
                unix_cmd_list = [tp_data['svn_template_cmd'] % {
                                 'subcommand': 'commit -m "%s" "%s"' %
                                 (_('ticket:%(id)s (on behalf of %(user)s)', id=tp_data['id'], user=author), tf.path)}]

                # Effective application of the list of commands
                retcode, lines = util.unix_cmd_apply(self.env, unix_cmd_list,
                                                     util.lineno())

                revision = ''
                if retcode == 0:
                    for line in reversed(lines):
                        if line.startswith(u'Révision '):
                            regular_expression = u'\ARévision (\d+) propagée\.\n\Z'
                            match = re.search(regular_expression, line)
                            if match:
                                revision = match.group(1)
                            break
                else:
                    raise TracError("[Commit of the ticket form]")

                #
                # Post-processing associated with
                # form or TRAC data modification
                #

                # Copy the updated form into its copy under .trac
                if not os.access(tf.copycontent_path, os.F_OK):
                    os.mkdir(tf.copycontent_path)
                if os.access(tf.copycontent_filename, os.F_OK):
                    os.remove(tf.copycontent_filename)
                shutil.copy(tf.content_filename, tf.copycontent_filename)

                # Update <TRAC to SVN> links
                if 'status' in old_values:
                    old_status = old_values['status']
                else:
                    old_status = None
                self._links_update(ticket,
                                   tp_data,
                                   revision,
                                   'change',
                                   'Ticket changed',
                                   old_status)

            # Any modification of attachments ?
            if os.access(tp_data['attachment'].destpath, os.F_OK):
                unix_cmd_list = [tp_data['svn_template_cmd'] % {'subcommand': 'st'} +
                                '"' +
                                tp_data['attachment'].destpath +
                                '" | grep "^M" | awk -F "      " \'{print $2}\'']

                # Effective application of the list of commands
                retcode, lines = util.unix_cmd_apply(self.env, unix_cmd_list,
                                                    util.lineno())

                if len(lines) != 0:  # at least one attachment has been modified
                    attachment_modified = True
                    dest_paths = [line.strip(' \n') for line in lines]
                    dest_filenames = [path[path.rfind('/') + 1:]
                                    for path in dest_paths]
                    attachments = dict(zip(dest_filenames, dest_paths))
                else:
                    attachment_modified = False
            else:
                attachment_modified = False

            if attachment_modified:

                # Update of the working copy -
                # attachments may have been added, removed or
                # edited by someone else (Force Edit mode)
                unix_cmd_list = [tp_data['svn_template_cmd'] % {
                                 'subcommand': 'up'} +
                                 '"' +
                                 tp_data['attachment'].destpath +
                                 '"']

                #
                # Commit of the attachments
                #
                unix_cmd_list += [tp_data['svn_template_cmd'] % {
                                  'subcommand': 'commit -m "%s" "%s"' % (_('ticket:%(id)s (on behalf of %(user)s)',
                                  id=tp_data['id'], user=author), tp_data['attachment'].destpath)}]

                # Effective application of the list of commands
                retcode, lines = util.unix_cmd_apply(self.env, unix_cmd_list,
                                                     util.lineno())

                revision = ''
                if retcode == 0:
                    for line in reversed(lines):
                        if line.startswith(u'Révision '):
                            regular_expression = u'\ARévision (\d+) propagée\.\n\Z'
                            match = re.search(regular_expression, line)
                            if match:
                                revision = match.group(1)
                            break
                else:
                    raise TracError("[Commit of the attachments]")

                #
                # Post-processing associated to attachments modification
                #

                # Copy of modified attachments to the Trac project environment
                for filename in dest_filenames:
                    a = Attachment(self.env,
                                   'ticket',
                                   ticket.id,
                                   filename)
                    shutil.copy(attachments[filename], a.path)

                # Update <TRAC to SVN> links
                self._links_update(ticket, tp_data,
                                   revision,
                                   'change',
                                   'Attachment(s) changed')

            if (not edit_form_modified and
                not trac_data_modified and
                not attachment_modified and
                'status' in old_values and
                old_values['status'] == '05-assigned_for_implementation' and
                ticket['status'] == '06-implemented'):
                regexp = ('\[\[BR\]\]%s \(%s\): '
                          '\[/browser.+?\?rev=(\d+) '
                          '.+? @ \d+\]' % (tp_data['ticket_type'],
                                           tp_data['ticket_form'].label))
                m = re.search(regexp, ticket['description'])
                if m:
                    revision = m.group(1)
                # Update <TRAC to SVN> links
                self._links_update(ticket,
                                   tp_data,
                                   revision,
                                   'change',
                                   'Document changed',
                                   old_values['status'])

            if ticket['type'] == 'MOM' and ticket['momtype'] in ('CCB', 'Review'):

                # Fill in milestone field from milestonetag field
                tg = model.Tag(self.env, ticket['milestonetag'])
                ticket['milestone'] = tg.tagged_item

                now = datetime.now(utc)
                # Beware recursion !
                ticket.save_changes('trac', 'Ticket changed', now)

        elif ticket['type'] == 'DOC' or (ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy) or ticket['type'] == 'FEE':

            if ticket['parent']:
                # update associated milestone tag
                pass

            template_cls = cache.Ticket_Cache.get_subclass(ticket['type'])
            with template_cls(self.env, self.trac_env_name,
                              authname, ticket) as doc:
                if ('status' in old_values and
                      old_values['status'] == TicketWF.get_WF(ticket).get_initial_status()):
                    # We are leaving the assigned_for_edition state
                    # Any change outside of Trac ?
                    last_rev_repo = util.get_last_path_rev_author(
                        self.env, util.get_url(ticket['sourceurl']))[2]
                    if last_rev_repo != util.get_revision(ticket['sourceurl']):
                        # The revision has changed
                        # Update ticket sourceurl automatically IF it does exist
                        new_sourceurl = '%s?rev=%s' % (util.get_url(ticket['sourceurl']), last_rev_repo)
                        new_url = util.get_repo_url(self.env, new_sourceurl)
                        if util.exist_in_repo(self.env, new_url):
                            ticket['sourceurl'] = new_sourceurl
                            now = datetime.now(utc)
                            ticket.save_changes('trac', _('Source Url changed (on behalf of %(user)s)', user=authname), now)
                if ('status' not in old_values and
                    ticket['status'] == TicketWF.get_WF(ticket).get_initial_status()):
                    # Changes on sourcefile or pdffile are taken into account
                    # when in the assigned_for_edition state
                    if 'sourcefile' in old_values and ticket['sourcefile'] != 'N/A':
                        doc.checkout(ticket['sourcefile'])
                        doc.add(ticket['sourcefile'])
                        if ticket['sourcefile'].endswith('.docm'):
                            doc.lock(ticket['sourcefile'])
                            doc.update_data(ticket['sourcefile'])
                            doc.upgrade_document(ticket['sourcefile'])
                    if 'pdffile' in old_values and ticket['pdffile'] != 'N/A':
                        doc.checkout(ticket['pdffile'])
                        doc.add(ticket['pdffile'])
                        if ticket['sourcefile'] and ticket['sourcefile'].endswith('.docm'):
                            doc.lock(ticket['pdffile'])
                    revision = doc.commit()
                    if revision != '':
                        # The locks have been automatically removed
                        # Beware recursion and cache semaphore !
                        ticket['sourceurl'] = '%s?rev=%s' % (
                            util.get_url(ticket['sourceurl']),
                            revision)
                        now = datetime.now(utc)
                        ticket.save_changes('trac', 'Ticket changed', now)
                    else:
                        # Remove the locks
                        if ('sourcefile' in old_values and
                            ticket['sourcefile'] != 'N/A' and
                            ticket['sourcefile'].endswith('.docm')):
                            doc.unlock(ticket['sourcefile'])
                        if ('pdffile' in old_values and
                            ticket['pdffile'] != 'N/A' and
                            ticket['sourcefile'] and
                            ticket['sourcefile'].endswith('.docm')):
                            doc.unlock(ticket['pdffile'])
                elif ((ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy) and
                      'status' in old_values and ticket['status'] == '05-assigned_for_sending'):
                    ticket['fromname'] = util.formatted_name(authname)
                    with Ldap_Utilities() as ldap_util:
                        ticket['fromemail'] = util.Users.get_email(self.env, authname, ldap_util)
                    ticket['fromphone'] = self.env.config.get('artusplugin', 'phone')
                    now = datetime.now(utc)
                    ticket.save_changes('trac', 'Ticket changed', now)
                elif (ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy and
                      (('status' in old_values and old_values['status'] == '05-assigned_for_sending' and
                        ticket['status'] == 'closed') or
                       ('status' not in old_values and ticket['status'] == '05-assigned_for_sending' and
                        ('owner' not in old_values or old_values['owner'] == authname)))):
                    # Fill in form fields (distribution)
                    doc.lock(ticket['pdffile'])
                    doc.fill_pdf(ticket['pdffile'])
                    revision = doc.commit()
                    if revision != '':
                        # The lock has been automatically released
                        # Beware recursion and cache semaphore !
                        ticket['sourceurl'] = '%s?rev=%s' % (
                            util.get_url(ticket['sourceurl']),
                            revision)
                        now = datetime.now(utc)
                        ticket.save_changes('trac', 'Ticket changed', now)
                    else:
                        # Remove the lock
                        doc.unlock(ticket['pdffile'])

        return

    def ticket_deleted(self, ticket):
        """Called when a ticket is deleted."""

        # authname is necessary for addressing the correct working copy
        req = util.get_req()
        authname = req.authname if req else 'trac'

        # removal of ticket via trac-admin IS supported

        tickets_with_forms = set(('EFR', 'ECR', 'RF', 'PRF'))

        if 'MOM' in util.get_prop_values(self.env,
                                         'ticket_edit.office_suite').keys():
            tickets_with_forms.add('MOM')

        if ticket['type'] in tickets_with_forms:
            tp_data = form.TicketForm.get_ticket_process_data(self.env, authname, ticket)
            tf = tp_data['ticket_form']
            if not util.exist_in_repo(self.env, tf.http_url):
                tickets_with_forms.remove(ticket['type'])

        if ticket['type'] in tickets_with_forms:
            # Subversion ticket backup is removed
            unix_cmd_list = [tp_data['svn_template_cmd'] % {
                             'subcommand': 'delete -m "Removal of ticket:' +
                             tp_data['id'] + '"'} + '"' + tf.http_url + '"']

            # Effective application of the list of commands
            retcode, lines = util.unix_cmd_apply(self.env, unix_cmd_list, util.lineno())

            if retcode != 0:
                message = tag.p("Removal of %s " % ticket['type'],
                                tag.em("%s" % ticket.id),
                                " backup in the repository has failed.",
                                class_="message")
                for line in lines:
                    message(tag.p(line))
                raise TracError(message)

            # All working copies associated with this ticket are deleted
            unix_cmd_list = ['find "%s" -name "t%s" -exec rm -Rf {} \; &> /dev/null' % (
                             tf.program_path, tf.id)]

            # Effective application of the list of commands
            util.unix_cmd_apply(self.env, unix_cmd_list, util.lineno())

        # Remove document entry
        if ticket['type'] == 'DOC':
            # The document entry has been added either when creating a tag
            # or when creating a DOC ticket
            # It is removed if there is no associated tag
            if not set(model.Tag.select(
                self.env,
                ['tracked_item = "%s"' % ticket['configurationitem']],
                tag_type='version_tags')):
                try:
                    # Delete
                    doc = model.Document(self.env, ticket['configurationitem'])
                    doc.delete()
                except ResourceNotFound:
                    pass

        if (ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy) or ticket['type'] == 'FEE' or ticket['type'] == 'DOC':
            # All working copies associated with this ticket are deleted
            unix_cmd_list = ['find "%s" -name "t%s" -exec rm -Rf {} \; &> /dev/null' % (
                             '/var/cache/trac/tickets/%s' % self.trac_env_name,
                             ticket.id)]

            # Effective application of the list of commands
            util.unix_cmd_apply(self.env, unix_cmd_list, util.lineno())

        # The ticket type chronological number is decremented
        if ticket['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(ticket).legacy:
            # chrono number not recycled
            number = None
        else:
            number = self._get_ticket_chrono_number(ticket)
        if number:
            now = datetime.now(localtz)
            ticket_chrono_filename = self._get_ticket_chrono_filename(
                ticket, now)
            util.del_ticket_number(ticket_chrono_filename, number)

        return

    def _get_ticket_chrono_filename(self, ticket, now):
        # Computes ticket chrono filename
        if ticket['type'] in ('EFR', 'ECM'):
            skills = self.env.config.get('ticket-custom', 'skill.options')
            if 'SYS' in skills:
                skill = ''
            else:
                skill = ticket['skill']
        elif ticket['type'] in ('ECR', 'RISK', 'AI', 'MEMO'):
            skill = ticket['skill']
        elif (ticket['type'] == 'MOM' and
              ((ticket['momtype'] == 'Review' and
                ticket['skill'] not in self._skills['proj']) or
               ticket['momtype'] == 'Audit')):
                skill = ticket['skill'] + now.strftime("%y")
        else:
            return None
        filename = '/srv/trac/%(trac_env_name)s/tickets/%(ticket_type)s/' \
                   '%(ticket_type)s%(skill)snb' \
                   % {'trac_env_name': self.trac_env_name,
                      'ticket_type': ticket['type'],
                      'skill': skill}
        return filename

    def _get_ticket_chrono_number(self, ticket):
        # Extract chrono number from ticket id
        if ticket['type'] in ('EFR', 'ECR', 'RISK', 'AI', 'MEMO'):
            try:
                chr_nb = int(ticket['summary'].rsplit('_', 1)[-1])
            except ValueError:
                chr_nb = None
        elif ticket['type'] in ('ECM'):
            try:
                chr_nb = int(ticket['summary'].split('_')[2])
            except ValueError:
                chr_nb = None
        elif (ticket['type'] == 'MOM' and
              ((ticket['momtype'] == 'Review' and
                ticket['skill'] not in self._skills['proj']) or
               ticket['momtype'] == 'Audit')):
            try:
                chr_nb = int(ticket['summary'].rsplit('-', 1)[-1])
            except TypeError:
                chr_nb = None
        else:
            chr_nb = None

        return chr_nb

    # IAttachmentChangeListener methods

    def attachment_added(self, attachment):
        """Called when an attachment is added."""

        if attachment.parent_realm == 'ticket':

            # Add entry in attachment_custom table if not already in it
            # It is handled in web_ui.py where the 'source_url' parameter
            # is known

            # authname is necessary for addressing
            # the correct working copy
            # the authname associated to the attachment gives
            # who last modified the attachment, not who is NOW
            # modifying the attachment
            req = util.get_req()
            authname = req.authname if req else 'trac'
            if authname == 'trac':
                # add of attachment via trac-admin NOT supported
                return
            self._attachment_handling(attachment, 'add', authname)

        return

    def attachment_deleted(self, attachment):
        """Called when an attachment is deleted."""

        if attachment.parent_realm == 'ticket':

            # Remove entry from attachment_custom table if can be found in it
            try:
                db = self.env.get_db_cnx()
                prop = model.AttachmentCustom(self.env,
                                              (attachment.parent_realm,
                                               attachment.parent_id,
                                               attachment.filename,
                                               'source_url'),
                                              db=db)
                prop.delete(db)
                db.commit()
            except ResourceNotFound:
                pass

            # authname is necessary for addressing
            # the correct working copy
            # the authname associated to the attachment gives
            # who last modified the attachment, not who is NOW
            # modifying the attachment
            req = util.get_req()
            authname = req.authname if req else 'trac'
            if authname == 'trac':
                # removal of attachment via trac-admin NOT supported
                return
            self._attachment_handling(attachment, 'delete', authname)

        return

    def _attachment_handling(self, attachment, operation, authname):
        """ The processing to be done
            when an attachment is added or removed """

        # Get the associated ticket (if any)
        if attachment.parent_realm == 'ticket':
            attachments = [a for a
                           in Attachment.select(self.env,
                                                attachment.parent_realm,
                                                attachment.parent_id)]
            env_filenames = [a.filename
                             for a
                             in attachments]

            ticket = Ticket(self.env, attachment.parent_id)
            tickets_with_forms = set(('EFR', 'ECR', 'RF', 'PRF'))

            if 'MOM' in util.get_prop_values(self.env,
                                             'ticket_edit.office_suite').keys():
                tickets_with_forms.add('MOM')

            if ticket['type'] in tickets_with_forms:
                tp_data = form.TicketForm.get_ticket_process_data(self.env, authname, ticket)
                tf = tp_data['ticket_form']
                if not util.exist_in_repo(self.env, tf.http_url):
                    tickets_with_forms.remove(ticket['type'])

            if ticket['type'] in tickets_with_forms:

                # Setup of working copy
                if not os.access(tf.type_subpath, os.F_OK):
                    os.makedirs(tf.type_subpath)

                # Have we MANUAL edits from the user ?
                if os.access(tf.copycontent_path, os.F_OK):
                    if filecmp.cmp(tf.content_filename,
                                   tf.copycontent_filename):
                        edit_form_modified = False
                    else:
                        edit_form_modified = True
                else:
                    edit_form_modified = False

                # Any update of the form in the repository ?
                revision = ''
                if os.access(tf.oldcontent_filename, os.F_OK):
                    # awk removes the error status
                    unix_cmd_list = [tp_data['svn_template_cmd'] % {
                                     'subcommand': 'st --show-updates'} +
                                     '"' +
                                     tf.oldcontent_filename +
                                     '" | grep \* | awk \'{print}\'']

                    # Effective application of the list of commands
                    retcode, lines = util.unix_cmd_apply(self.env,
                                                         unix_cmd_list,
                                                         util.lineno())

                    if len(lines) != 0:  # there is an update in the repository
                        repo_form_modified = True
                        # Get working revision before updating
                        # so as to be able to revert to it
                        m = re.search(r"\A\s*\*\s*(\d+)\s*/var/cache",
                                      lines[0])
                        if m:
                            revision = m.group(1)
                    else:
                        repo_form_modified = False
                else:
                    repo_form_modified = True

                if repo_form_modified:
                    # The ticket is checked out
                    unix_cmd_list = [tp_data['svn_template_cmd'] % {
                                     'subcommand': 'co --depth files'} +
                                     '"' +
                                     tf.http_url +
                                     '" "' +
                                     tf.path +
                                     '"']

                    # Effective application of the list of commands
                    util.unix_cmd_apply(self.env, unix_cmd_list, util.lineno())

                    if not edit_form_modified:
                        # Copy the updated form onto the edited one
                        # ONLY if the Ooo form carried NO user modifications
                        if not os.access(tf.content_path, os.F_OK):
                            os.mkdir(tf.content_path)
                        shutil.copy(tf.oldcontent_filename,
                                    tf.content_filename)

                if ticket.time_created > tf.get_update_deployment_date():
                    # Note: If the Ooo form has changes when an attachment
                    # is added or removed, they are merged
                    # with the attachment data
                    tf.update(ticket, tp_data, ['attachments', None], None)

                if edit_form_modified and repo_form_modified:
                    # Revert to the preceding revision
                    # so a conflict is detected and handled by the user
                    if revision:
                        unix_cmd_list = [tp_data['svn_template_cmd'] % {
                                         'subcommand': 'up -r ' +
                                         revision} +
                                         '"' +
                                         tf.oldcontent_filename +
                                         '"']

                        # Effective application of the list of commands
                        util.unix_cmd_apply(self.env, unix_cmd_list, util.lineno())
                else:
                    if not edit_form_modified:
                        # Copy the updated form onto the edited one
                        # ONLY if the Ooo form carried NO user modifications
                        if not os.access(tf.content_path, os.F_OK):
                            os.mkdir(tf.content_path)
                        shutil.copy(tf.oldcontent_filename,
                                    tf.content_filename)

                        #
                        # Commit of the ticket form
                        #
                        unix_cmd_list = [tp_data['svn_template_cmd'] % {
                                         'subcommand': 'commit -m "%s" "%s"' %
                                         (_('ticket:%(id)s (on behalf of %(user)s)', id=tp_data['id'], user=authname), tf.path)}]

                        # Effective application of the list of commands
                        retcode, lines = util.unix_cmd_apply(self.env, unix_cmd_list,
                                                             util.lineno())

                        revision = ''
                        if retcode == 0:
                            for line in reversed(lines):
                                if line.startswith(u'Révision '):
                                    regular_expression = u'\ARévision (\d+) propagée\.\n\Z'
                                    match = re.search(regular_expression, line)
                                    if match:
                                        revision = match.group(1)
                                    break
                        else:
                            raise TracError("[Commit of the ticket form]")

                        #
                        # Post-processing associated with
                        # form or TRAC data modification
                        #

                        # Copy the updated form into its copy under .trac
                        if not os.access(tf.copycontent_path, os.F_OK):
                            os.mkdir(tf.copycontent_path)
                        shutil.copy(tf.content_filename, tf.copycontent_filename)

                        # Update <TRAC to SVN> links
                        old_status = None
                        self._links_update(ticket,
                                           tp_data,
                                           revision,
                                           'change',
                                           'Ticket changed',
                                           old_status)

                # Extraction of 'attachment url' folder (in case)
                unix_cmd_list = ['if [ ! -d "' +
                                 tp_data['attachment'].destpath +
                                 '" ]; then ' +
                                 tp_data['svn_template_cmd'] % {
                                     'subcommand': 'co --depth files'} +
                                 '"' +
                                 tp_data['attachment'].http_url +
                                 '" "' +
                                 tp_data['attachment'].destpath +
                                 '"; fi']

                # Update of attachments working copy sub-tree
                # before any modification of the wc tree -
                # attachments may have been added or removed by someone else
                unix_cmd_list += [tp_data['svn_template_cmd'] % {
                                  'subcommand': 'up'} +
                                  '"' +
                                  tp_data['attachment'].destpath +
                                  '"']

                # Effective application of the list of commands
                util.unix_cmd_apply(self.env, unix_cmd_list, util.lineno())

                # Checklist header update
                if (operation == 'add' and
                    ticket['type'] == 'RF' and
                    attachment.filename.startswith('CHKLST') and
                    (attachment.filename ==
                     ticket['summary'].replace('RF_', 'CHKLST_') +
                     '.' +
                     attachment.filename.rsplit('.', 1)[-1])):
                    checker = ticket['summary'].rsplit('_', 1)[-1]
                    checklist = form.CheckList(self.env,
                                               ticket,
                                               tp_data,
                                               attachment,
                                               checker)
                    checklist.update()

                # Backup of attachment modifications
                cache_filenames = [filepath.replace(
                                   tp_data['attachment'].destpath.encode('utf-8') + '/', '')
                                   for filepath in
                                   glob.glob(tp_data['attachment'].destpath.encode('utf-8') +
                                             '/*')
                                   ]

                if cache_filenames:
                    shutil.rmtree(tp_data['attachment'].destpath_bkup, ignore_errors=True)
                    shutil.copytree(tp_data['attachment'].destpath, tp_data['attachment'].destpath_bkup)

                    # Synchronizing of attachments (if any) -
                    # except modifications which are only pushed to TRAC env
                    # and subversion through the ticket 'Submit changes'
                    for f in glob.glob(tp_data['attachment'].destpath.encode('utf-8') + '/*'):
                        os.remove(f)

                # Synchronization of attachments between TRAC env and cache
                for a in attachments:
                    shutil.copy2(a.path, '%s/%s' % (tp_data['attachment'].destpath, a.filename))

                # Add or remove from subversion
                if operation == 'add':
                    unix_cmd_list = [tp_data['svn_template_cmd'] % {'subcommand': 'st'} +
                                     '"' +
                                     tp_data['attachment'].destpath +
                                     '" | grep "^\?" | sed -e \'s/^? */"/\' |'
                                     ' sed -e \'s/$/@"/\' |'
                                     ' xargs --no-run-if-empty svn add'
                                     ]
                else:
                    unix_cmd_list = [tp_data['svn_template_cmd'] % {
                                     'subcommand': 'st'} +
                                     '"' +
                                     tp_data['attachment'].destpath +
                                     '" | grep "^\!" | sed -e \'s/^! */"/\' |'
                                     ' sed -e \'s/$/@"/\' |'
                                     ' xargs --no-run-if-empty svn delete']

                # Commit of added or removed attachments
                unix_cmd_list += [tp_data['svn_template_cmd'] % {
                                  'subcommand': 'commit -m "%s" --depth files "%s"' %
                                  (_('ticket:%(id)s (on behalf of %(user)s)', id=tp_data['id'], user=authname),
                                   tp_data['attachment'].destpath)}]

                # Effective application of the list of commands
                retcode, lines = util.unix_cmd_apply(self.env, unix_cmd_list,
                                                     util.lineno())

                revision = ''
                if retcode == 0:
                    for line in reversed(lines):
                        if line.startswith(u'Révision '):
                            regular_expression = u'\ARévision (\d+) propagée\.\n\Z'
                            match = re.search(regular_expression, line)
                            if match:
                                revision = match.group(1)
                            break
                else:
                    raise TracError("[Commit of added or removed attachments]")

                # Restore of modifications
                if cache_filenames:
                    for env_filename in env_filenames:
                        if env_filename in cache_filenames:
                            shutil.copy2('%s/%s' % (tp_data['attachment'].destpath_bkup, env_filename),
                                         tp_data['attachment'].destpath)
                    shutil.rmtree(tp_data['attachment'].destpath_bkup, ignore_errors=True)

                # Update <TRAC to SVN> links
                if operation == 'add':
                    comment = 'Attachment added: '
                else:
                    comment = 'Attachment deleted: '
                self._links_update(ticket,
                                   tp_data,
                                   revision,
                                   'change',
                                   comment + attachment.filename)

        return

    def _links_update(self, ticket, data, revision, mode, comment, old_status=None):
        """Updates the TRAC to SVN links
           only in case of effective changes to commit """

        if revision != '':

            template_link = ''

            # Try to get back the template link
            if ticket['description']:
                end_idx = ticket['description'].find('[[BR]]')
                if end_idx != -1:
                    link = ticket['description'][:end_idx]
                    regexp = '%s template' % data['ticket_type']
                    if re.search(regexp, link):
                        template_link = link

            # No template link: new or old ticket
            if not template_link:
                template_link = data['ticket_form_template'].get_link()

            ticket_link = '[[BR]]%s (%s): [/browser%s/%s?rev=%s %s @ %s]' % (
                data['ticket_type'],
                data['ticket_form'].label,
                data['ticket_form'].repo_subpath.replace(' ', '%20'),
                data['ticket_id'].replace(' ', '%20'),
                revision,
                data['ticket_id'],
                revision)

            if data['ticket_type'] in ('RF', 'PRF'):
                # Document links
                document_links = self._get_document_links(ticket,
                                                          data,
                                                          mode,
                                                          old_status)
            else:
                document_links = ''

            ticket['description'] = '%s%s%s' % (template_link,
                                                ticket_link,
                                                document_links)
            now = datetime.now(utc)
            ticket.save_changes('trac', comment, now)

    def _get_document_links(self, ticket, data, mode, old_status):
        """ Return links on the document:
                under 'tags' for the original version
                under 'trunk' for the modified version
            Both are inputs of the verification process
            Also UPDATES the ticket['documenturl'] with the current tags rev """

        # Document links
        document_links = ''

        if ticket['documenturl'] != "":

            if mode == 'creation':
                # Revision under review
                tagsrev = util.get_last_path_rev_author(self.env, ticket['documenturl'])[2]

                # Get the document link on the last verified version by the logged in user
                # This if for comparing two versions or status therefore it will be available
                # only if the PRF is associated to a DOC/ECM/FEE ticket
                if ticket['parent']:
                    # Parent of PRF ticket should be a DOC/ECM/FEE ticket
                    tid = int(ticket['parent'].lstrip('#'))
                    tkt = Ticket(self.env, tid)
                    if not ((tkt['type'] == 'ECM' and not web_ui.Ticket_UI.get_UI(tkt).legacy and
                             tkt['ecmtype'] == 'Technical Note') or
                             tkt['type'] == 'FEE' or
                             tkt['type'] == 'DOC'):
                        raise TracError(tag.p("Parent ticket ",
                                              tag.a("#%s" % tid, href="%s/ticket/%s" % (data['base_path'], tid)),
                                              " is not a DOC ticket or an ECM Technical Note ticket or a FEE ticket"))
                    ci_name = tkt['configurationitem']
                    # Get all DOC tickets on the same CI
                    db = self.env.get_db_cnx()
                    cursor = db.cursor()
                    cursor.execute("SELECT t.id FROM ticket t,ticket_custom tc "
                                   "WHERE t.type='%s' "
                                   "AND t.id=tc.ticket "
                                   "AND tc.name='configurationitem' "
                                   "AND tc.value='%s'" % (tkt['type'], ci_name))
                    parent_tkts = [row[0] for row in cursor]
                    prf_tkts = []
                    for tid in parent_tkts:
                        # For all parent tickets, get logged in user closed and fixed PRF tickets on them
                        cursor.execute("SELECT t.id FROM ticket t,ticket_custom tc "
                                       "WHERE t.id=tc.ticket "
                                       "AND t.type='PRF' "
                                       "AND t.summary LIKE '%%_%s' "
                                       "AND t.status='closed' "
                                       "AND t.resolution='fixed' "
                                       "AND tc.name='parent' "
                                       "AND tc.value='#%s'" % (ticket['owner'], tid))
                        prf_tkts.extend([row[0] for row in cursor])
                    prf_revs_links = []
                    # Get revision and wiki link on each version or status of the document that the logged in user reviewed - modified or not
                    for tid in prf_tkts:
                        tkt = Ticket(self.env, tid)
                        description = tkt['description']
                        if description:
                            document_index = description.rfind('[[BR]]Document')
                            document_line = description[document_index:]
                            wiki_index = document_line.rfind('[')
                            wiki_link = document_line[wiki_index:]
                            m = re.search(r'\?rev=(\d+)', wiki_link)
                            if m:
                                prf_revs_links.append((tid, m.group(1), wiki_link))

                    # Get the logged in user last PRF on the document and its revision
                    if prf_revs_links:
                        prf_revs_links.sort(key=lambda x: int(x[1]))
                        last_rev = prf_revs_links[-1][1]
                    else:
                        last_rev = None

                    # If that revision is different from the one under review (document has been modified)
                    # then display it under Subversion links
                    if last_rev and int(last_rev) < int(tagsrev):
                        tid = prf_revs_links[-1][0]
                        wiki_link = prf_revs_links[-1][2]
                        document_links = "[[BR]]Document (previously reviewed through #%s): %s" % (tid, wiki_link)
                    else:
                        document_links = "[[BR]]Document (not previously reviewed)"

                # Get the document link on the original version -
                #  at this stage it has no revision (so HEAD revision)
                tagurl_address = ticket['documenturl']
                tagurl_address = tagurl_address.replace(' ', '%20')
                tagurl_text = tagurl_address[tagurl_address.rfind("/") + 1:]
                tagurl_text = tagurl_text.replace('%20', ' ')
                # No ticket may be created when browsing with a revision
                # so this is HEAD revision and we want the tag creation revision
                document_links += ('[[BR]]Document (reviewed):'
                                   ' [/browser%s?rev=%s %s @ %s]' % (
                                       tagurl_address,
                                       tagsrev,
                                       tagurl_text,
                                       tagsrev)
                                  )
                # UPDATE
                ticket['documenturl'] += '@' + tagsrev
            else:
                # Get back the document links
                if ticket['description']:
                    document_index = ticket['description'].find('[[BR]]Document')
                    if document_index != -1:
                        document_links = ticket['description'][document_index:]

                        # The document modified version is updated
                        # at each implementation transition
                        if (old_status and
                            old_status == '05-assigned_for_implementation' and
                            ticket['status'] == '06-implemented'):
                            modified_index = ticket['description'].rfind('[[BR]]Document (modified)')

                            # modified version updated ?
                            if modified_index != -1:
                                document_links = ticket['description'][document_index:modified_index]
                            else:
                                # Legacy
                                modified_index = ticket['description'].rfind('[[BR]]Document (Verification modified)')
                                # modified version updated ?
                                if modified_index != -1:
                                    document_links = ticket['description'][document_index:modified_index]

                            # Get the document url and revision of the original document in the trunk
                            trunkurl, trunkrev = util.get_trunk_url_rev_from_tag(self.env, ticket)

                            # Get the document url and revision of the modified document in the trunk
                            # We want the last revision even through path moves
                            last_path_rev_author = util.get_last_path_rev_author(self.env, trunkurl, trunkrev)
                            reponame = last_path_rev_author[0]
                            trunkurl = last_path_rev_author[1]
                            trunkrev = last_path_rev_author[2]

                            if trunkurl != '' and trunkrev != '':
                                if reponame:
                                    trunkurl = '/' + reponame + trunkurl
                                trunkurl_address = trunkurl.replace(' ', '%20')
                                trunkurl_text = trunkurl[trunkurl.rfind("/") + 1:]
                                trunkurl_text = trunkurl_text.replace('%20', ' ')
                                document_links += '[[BR]]Document (modified): [/browser%s?rev=%s %s @ %s]' % (
                                    trunkurl_address,
                                    trunkrev,
                                    trunkurl_text,
                                    trunkrev)
                            else:
                                document_links += ('[[BR]]Document (modified): '
                                                   'The document has not been found, it has probably been relocated')

        return document_links
